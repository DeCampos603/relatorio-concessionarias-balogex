# -*- coding: utf-8 -*-
"""
Gera o Relatório Consolidado de Concessionárias (água e energia) da BA Ap Log Ex
e das OMDS apoiadas, a partir dos arquivos .xls exportados do SAG.

Entrada : G:\\Meu Drive\\REPO\\RELATORIO CONCESSIONARIAS  (árvore já organizada)
Saída   : site/index.html (autocontido) + site/dados.json

Uso:
    py -3 gerar_site.py
    py -3 gerar_site.py --fonte "OUTRO\\CAMINHO"
"""
import os, re, sys, json, html, argparse, datetime, collections

AQUI = os.path.dirname(os.path.abspath(__file__))
FONTE_PADRAO = r"G:\Meu Drive\REPO\RELATORIO CONCESSIONARIAS"
SAIDA = os.path.join(AQUI, "site")

MES = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
ANOS = ["2023", "2024", "2025", "2026"]

# UG, sigla, RM, comando, CGCFEx e UF vêm dos próprios arquivos do SAG.
# A denominação por extenso ("nome") NÃO é exportada pelo sistema: foi informada pela
# BA Ap Log Ex. O nome da concessionária continua sem fonte em todo o acervo.
UGS = [
    {"ug": "160238", "sigla": "Ba Ap Log Ex", "slug": "BA-AP-LOG-EX", "logo": "BAAPLOG.png",
     "nome": "Base de Apoio Logístico"},
    {"ug": "160246", "sigla": "DC Mun",       "slug": "DC-MUN",       "logo": "DCMUN.png",
     "nome": "Depósito Central de Munição"},
    {"ug": "160304", "sigla": "BMSA",         "slug": "BMSA",         "logo": "BMSA.png",
     "nome": "Batalhão de Manutenção e Suprimento de Armamento"},
    {"ug": "160307", "sigla": "1º D Sup",     "slug": "1-D-SUP",      "logo": "1DSUP.png",
     "nome": "1º Depósito de Suprimentos"},
    {"ug": "160321", "sigla": "ECT",          "slug": "ECT",          "logo": "Ect.png",
     "nome": "Estabelecimento Central de Transportes"},
    {"ug": "160329", "sigla": "BCMS",         "slug": "BCMS",         "logo": "BCMS.png",
     "nome": "Batalhão Central de Manutenção e Suprimento"},
]
UGIDX = {u["ug"]: u for u in UGS}

METRICA_POR_CAPTION = {
    "CONSUMO EM KW/H NA PONTA":      "kwh_ponta",
    "CONSUMO EM KW/H FORA DA PONTA": "kwh_fora",
    "VALOR EM R$ NA PONTA":          "rs_ponta",
    "VALOR EM R$ FORA DA PONTA":     "rs_fora",
    "CONSUMO EM M3":                 "m3",
    "VALOR EM R$ DA FATURA":         "rs_agua",
}

# Abaixo deste valor total anual em R$, para centenas de MWh consumidos, o campo
# "valor" só pode conter tarifa unitária, não fatura.
LIMIAR_TARIFA = 50.0

# --- Execução orçamentária (crosstabs do Tesouro Gerencial) -------------------
# Planos Internos das concessionárias, na ordem em que aparecem no relatório.
PIS = [
    ("I3DACSPAGES", "Água e esgoto",     "agua"),
    ("I3DACSPENEL", "Energia elétrica",  "energia"),
    ("I3DACSPTELM", "Telefonia móvel",   "telm"),
    ("I3DACSPTELF", "Telefonia fixa",    "telf"),
    ("I3DACSPCORR", "Serviços postais",  "corr"),
]
PI_ROT = {p: r for p, r, _ in PIS}
METRICAS_ORC = ["PROVISAO RECEBIDA", "DESPESAS EMPENHADAS", "DESPESAS LIQUIDADAS",
                "DESPESAS PAGAS", "PAGAMENTOS TOTAIS (EXERCICIO + RP)",
                "RESTOS A PAGAR NAO PROCESSADOS INSCRITOS",
                "RESTOS A PAGAR NAO PROCESSADOS REINSCRITOS",
                "RESTOS A PAGAR NAO PROCESSADOS CANCELADOS",
                "RESTOS A PAGAR NAO PROCESSADOS LIQUIDADOS",
                "RESTOS A PAGAR NAO PROCESSADOS PAGOS",
                "RESTOS A PAGAR PROCESSADOS PAGOS"]
VAZIOS_TG = ("-9", "'-9", "-8", "'-8", "", "None")
PERIODO_TG = re.compile(r"^([A-Z]{3})/(\d{4})$")
PREFIXOS_METRICA = ("DESPESAS", "RESTOS A PAGAR", "PAGAMENTOS", "PROVISAO", "CREDITO",
                    "LIQUIDACOES", "RAP", "VALORES", "RECEITA")
MES_TG = {"JAN": 1, "FEV": 2, "MAR": 3, "ABR": 4, "MAI": 5, "JUN": 6,
          "JUL": 7, "AGO": 8, "SET": 9, "OUT": 10, "NOV": 11, "DEZ": 12}


# ------------------------------------------------------------------------ leitura

def ler(caminho):
    with open(caminho, "rb") as f:
        b = f.read()
    if b[:3] == b"\xef\xbb\xbf":
        b = b[3:]
    return b.decode("utf-8", errors="replace")


def texto(s):
    return html.unescape(re.sub(r"<[^>]+>", "", s)).replace("\u00a0", " ").strip()


def numero(s):
    s = texto(s)
    if s in ("", "-"):
        return None
    try:
        return float(s.replace(".", "").replace(",", "."))
    except ValueError:
        return None


def linhas(t):
    return [re.findall(r"<t[hd][^>]*>(.*?)</t[hd]>", tr, re.S)
            for tr in re.findall(r"<tr[^>]*>(.*?)</tr>", t, re.S)]


# ------------------------------------------------------------------------ coleta

def coletar(fonte):
    """Percorre a árvore e devolve séries, metas por UG, consolidados e inventário."""
    series = {u["ug"]: {} for u in UGS}
    meta_agua = {u["ug"]: None for u in UGS}
    meta_energia = {u["ug"]: {} for u in UGS}
    consolidado = {}
    inventario = []
    pastas = set()

    for raiz, dirs, arquivos in os.walk(fonte):
        for d in dirs:
            pastas.add(os.path.join(raiz, d))
        for fn in sorted(arquivos):
            if not fn.lower().endswith(".xls"):
                continue
            caminho = os.path.join(raiz, fn)
            rel = os.path.relpath(caminho, fonte)
            t = ler(caminho)
            m = re.search(r"<caption[^>]*>(.*?)</caption>", t, re.S)
            cap = texto(m.group(1)) if m else ""
            rs = linhas(t)

            if " DA UG: " in cap:                        # série histórica 2023–2026
                metrica, resto = cap.split(" DA UG: ", 1)
                ug = resto[:6]
                chave = METRICA_POR_CAPTION[metrica.strip()]
                anos = [texto(c) for c in rs[0]][1:]
                vals = {a: [None] * 12 for a in anos}
                for r in rs[1:]:
                    if not r:
                        continue
                    i = MES.index(texto(r[0]))
                    for j, a in enumerate(anos):
                        vals[a][i] = numero(r[1 + j])
                series[ug][chave] = vals
                inventario.append({"arquivo": rel, "ug": ug, "conteudo": cap})

            elif cap.startswith("CONSUMO DE ÁGUA"):      # consumo × meta 2026, água
                ug = re.search(r"UG: (\d{6})", cap).group(1)
                consumo, alvo = [None] * 12, [None] * 12
                for r in rs[1:]:
                    if not r:
                        continue
                    i = MES.index(texto(r[0]))
                    consumo[i] = numero(r[1])
                    alvo[i] = numero(r[2]) if len(r) > 2 else None
                meta_agua[ug] = {"consumo": consumo, "meta": alvo}
                inventario.append({"arquivo": rel, "ug": ug, "conteudo": cap})

            elif "CONSUMO DE ENERGIA" in cap:            # consumo × meta 2026, energia
                mm = re.search(r"PER[IÍ]ODO\s+(FORA PONTA|PONTA)\s*-\s*UG:\s*(\d{6})", cap)
                posto = "fora" if mm.group(1) == "FORA PONTA" else "ponta"
                ug = mm.group(2)
                consumo, alvo = [None] * 12, [None] * 12
                for r in rs[1:]:
                    if not r:
                        continue
                    i = MES.index(texto(r[0]))
                    consumo[i] = numero(r[1])
                    alvo[i] = numero(r[2]) if len(r) > 2 else None
                meta_energia[ug][posto] = {"consumo": consumo, "meta": alvo}
                inventario.append({"arquivo": rel, "ug": ug, "conteudo": cap})

            else:                                        # relatório consolidado de meta
                servico = "agua" if "AGUA" in fn.upper() else "energia"
                cab = [texto(c) for c in rs[0]]
                dados = []
                for r in rs[1:]:
                    if not r:
                        continue
                    dados.append({"ug": texto(r[0]), "sigla": texto(r[1]),
                                  "mes": int(numero(r[6])), "valores": [numero(c) for c in r[7:]]})
                consolidado[servico] = {"colunas": cab[7:], "dados": dados}
                inventario.append({"arquivo": rel, "ug": "TODAS",
                                   "conteudo": f"Relatório consolidado de meta × consumo ({servico})"})

    return series, meta_agua, meta_energia, consolidado, inventario, len(pastas)


def coletar_orcamento(fonte):
    """Lê os crosstabs 'CRÉDITO DISP*.xlsx' do Tesouro Gerencial.

    O layout varia entre extrações: em umas a métrica está na linha 6 e o período
    (DEZ/aaaa) na 7; em outras o período está na 5 e a métrica na 7. As duas linhas são
    detectadas pelo conteúdo, não pela posição.

    As colunas A a M identificam UG executora, Plano Interno e natureza de despesa.
    Linhas com natureza vazia são subtotais do próprio crosstab e são descartadas —
    sem isso o valor dobraria.
    """
    try:
        import warnings
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
        import openpyxl
    except ImportError:
        print("AVISO: openpyxl não instalado — bloco orçamentário não será gerado.\n"
              "       pip install openpyxl")
        return None, []

    arquivos = sorted(f for f in os.listdir(fonte)
                      if f.lower().endswith(".xlsx") and "CR" in f.upper() and "DISP" in f.upper())
    if not arquivos:
        return None, []

    reg = collections.defaultdict(float)       # (ano, ug, pi, metrica) -> R$
    total_ug = collections.defaultdict(float)  # (ano, ug, metrica) -> R$ (todos os PI)
    ugs_validas = {u["ug"] for u in UGS}
    lidos, corte = [], {}

    def limpo(v):
        s = str(v).strip() if v is not None else ""
        return "" if s in VAZIOS_TG else s

    for fn in arquivos:
        wb = openpyxl.load_workbook(os.path.join(fonte, fn), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        cab = [r for i, r in enumerate(ws.iter_rows(max_row=8, values_only=True))]

        lin_per = lin_met = None
        for i in range(4, min(8, len(cab))):
            celulas = [str(v).strip() for v in cab[i][13:] if v]
            if any(PERIODO_TG.match(x) for x in celulas):
                lin_per = i
            if any(x.startswith(PREFIXOS_METRICA) for x in celulas):
                lin_met = i
        if lin_per is None or lin_met is None:
            print(f"AVISO: layout não reconhecido em {fn} — arquivo ignorado.")
            wb.close()
            continue

        colunas, metrica, periodo = {}, None, None
        for c in range(13, ws.max_column):
            m = cab[lin_met][c] if c < len(cab[lin_met]) else None
            p = cab[lin_per][c] if c < len(cab[lin_per]) else None
            if m:
                metrica = str(m).strip()
            if p:
                periodo = str(p).strip()
            if metrica and periodo:
                mm = PERIODO_TG.match(periodo)
                if mm:
                    colunas[c] = (metrica, mm.group(2), mm.group(1))

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 8:
                continue
            ug, pi, nat = limpo(row[2]), limpo(row[6]), limpo(row[8])
            if ug not in ugs_validas or not nat:
                continue
            for c, (met, ano, mes) in colunas.items():
                corte[ano] = mes
                if met not in METRICAS_ORC:
                    continue
                v = row[c] if c < len(row) else None
                if not isinstance(v, (int, float)) or not v:
                    continue
                total_ug[(ano, ug, met)] += v
                if pi in PI_ROT:
                    reg[(ano, ug, pi, met)] += v
        wb.close()
        lidos.append(fn)

    anos = sorted({k[0] for k in reg})
    periodos = {a: {"ano": a, "mes": corte.get(a, "DEZ"),
                    "fechado": corte.get(a, "DEZ") == "DEZ",
                    "n_meses": MES_TG.get(corte.get(a, "DEZ"), 12),
                    "rotulo": a if corte.get(a, "DEZ") == "DEZ"
                              else f"{a} até {corte[a].lower()}"} for a in anos}
    return {"reg": reg, "total_ug": total_ug, "anos": anos, "periodos": periodos}, lidos


# --------------------------------------------------------------------- utilidades

def soma(v, n=12):
    return sum(x for x in v[:n] if x)


def meses_com_dado(v):
    return [MES[i] for i in range(12) if v[i]]


def var(a, b):
    return None if not a else (b / a - 1) * 100


def fmt(v, casas=0):
    """Número em pt-BR: milhar com ponto, decimal com vírgula."""
    if v is None:
        return "—"
    return f"{v:,.{casas}f}".replace(",", "\u0001").replace(".", ",").replace("\u0001", ".")


def pc(v, casas=1):
    """Percentual em pt-BR, sempre com sinal."""
    return "—" if v is None else ("+" if v >= 0 else "") + fmt(v, casas) + "%"


# -------------------------------------------------------------------- indicadores

def detecta_tarifario(series):
    """UGs em que o campo VALOR do SAG contém a TARIFA (R$/kWh), não a fatura."""
    fora = []
    for u in UGS:
        s = series[u["ug"]]
        rs = soma(s["rs_ponta"]["2025"]) + soma(s["rs_fora"]["2025"])
        kwh = soma(s["kwh_ponta"]["2025"]) + soma(s["kwh_fora"]["2025"])
        if kwh > 10000 and rs < LIMIAR_TARIFA:
            fora.append(u["ug"])
    return fora


def bases_energia(series):
    """Meses de 2026 com lançamento em cada UG — a janela de comparação daquela unidade."""
    return {u["ug"]: len([1 for i in range(6)
                          if (series[u["ug"]]["kwh_ponta"]["2026"][i] or 0)
                          + (series[u["ug"]]["kwh_fora"]["2026"][i] or 0)])
            for u in UGS}


def energia_valores(series, tarifarias, bases):
    """Valor de energia por UG/ano, reconstruído onde for preciso.

    Usa a MESMA janela de meses que o consumo daquela UG (`bases`); do contrário a
    tabela mostraria R$ de seis meses ao lado de kWh de cinco.
    """
    out = {}
    for u in UGS:
        ug = u["ug"]
        s = series[ug]
        n = bases[ug]
        recon = ug in tarifarias
        por_ano = {}
        for a in ANOS:
            if recon:
                por_ano[a] = sum((s["kwh_ponta"][a][i] or 0) * (s["rs_ponta"][a][i] or 0) +
                                 (s["kwh_fora"][a][i] or 0) * (s["rs_fora"][a][i] or 0)
                                 for i in range(n))
            else:
                por_ano[a] = soma(s["rs_ponta"][a], n) + soma(s["rs_fora"][a], n)
        out[ug] = {"reconstruido": recon, "valores": por_ano, "meses": n}
    return out


def decomposicao_agua(series, n=6):
    """Separa a variação 2025→2026 do gasto com água em efeito preço e efeito volume."""
    itens, tot = [], {"m25": 0, "m26": 0, "r25": 0, "r26": 0}
    for u in UGS:
        s = series[u["ug"]]
        m25, r25 = soma(s["m3"]["2025"], n), soma(s["rs_agua"]["2025"], n)
        m26, r26 = soma(s["m3"]["2026"], n), soma(s["rs_agua"]["2026"], n)
        p25, p26 = r25 / m25, r26 / m26
        itens.append({
            "ug": u["ug"], "sigla": u["sigla"], "slug": u["slug"], "logo": u["logo"],
            "m25": m25, "m26": m26, "r25": r25, "r26": r26, "p25": p25, "p26": p26,
            "var_m": var(m25, m26), "var_r": var(r25, r26), "var_p": var(p25, p26),
            "efeito_preco": (p26 - p25) * m25,
            "efeito_volume": (m26 - m25) * p25,
            "efeito_misto": (p26 - p25) * (m26 - m25),
        })
        for k, v in (("m25", m25), ("m26", m26), ("r25", r25), ("r26", r26)):
            tot[k] += v
    p25, p26 = tot["r25"] / tot["m25"], tot["r26"] / tot["m26"]
    tot.update({
        "p25": p25, "p26": p26,
        "var_m": var(tot["m25"], tot["m26"]), "var_r": var(tot["r25"], tot["r26"]),
        "var_p": var(p25, p26),
        "efeito_preco": (p26 - p25) * tot["m25"],
        "efeito_volume": (tot["m26"] - tot["m25"]) * p25,
        "efeito_misto": (p26 - p25) * (tot["m26"] - tot["m25"]),
    })
    return {"itens": itens, "total": tot}


def serie_tarifa_agua(series, n=6):
    out = []
    for a in ANOS:
        m = sum(soma(series[u["ug"]]["m3"][a], n) for u in UGS)
        r = sum(soma(series[u["ug"]]["rs_agua"][a], n) for u in UGS)
        out.append({"ano": a, "m3": m, "rs": r, "tarifa": r / m})
    return out


def metas(consolidado, meta_energia):
    """Acumulado × meta por UG. Água vem do consolidado; energia, dos gráficos por posto."""
    res = {"agua": [], "energia": []}

    d = consolidado["agua"]["dados"]
    for u in UGS:
        regs = [r for r in d if r["ug"] == u["ug"]]
        if not regs:
            continue
        r = max(regs, key=lambda x: x["mes"])
        acum, alvo = r["valores"][6], r["valores"][7]
        res["agua"].append({"ug": u["ug"], "sigla": u["sigla"], "logo": u["logo"],
                            "mes": r["mes"], "acumulado": acum, "meta": alvo,
                            "desvio": var(alvo, acum)})

    for u in UGS:
        me = meta_energia[u["ug"]]
        if not me:
            continue
        reg = {"ug": u["ug"], "sigla": u["sigla"], "logo": u["logo"]}
        n = 0
        for posto in ("ponta", "fora"):
            c = me[posto]["consumo"]
            n = max(n, len(meses_com_dado(c)))
        reg["mes"] = n
        for posto in ("ponta", "fora"):
            c, alvo = me[posto]["consumo"], me[posto]["meta"]
            reg[f"{posto}_ac"] = soma(c, n)
            reg[f"{posto}_meta"] = soma(alvo, n)
            reg[f"{posto}_desvio"] = var(reg[f"{posto}_meta"], reg[f"{posto}_ac"])
            reg[f"{posto}_meta_ano"] = soma(alvo, 12)
        reg["total_ac"] = reg["ponta_ac"] + reg["fora_ac"]
        reg["total_meta"] = reg["ponta_meta"] + reg["fora_meta"]
        reg["desvio"] = var(reg["total_meta"], reg["total_ac"])
        res["energia"].append(reg)

    res["fora_agua"] = sum(1 for m in res["agua"] if m["desvio"] and m["desvio"] > 0)
    res["fora_energia_fora"] = sum(1 for m in res["energia"]
                                   if m["fora_desvio"] and m["fora_desvio"] > 0)
    return res


def orcamento(orc, series, tarifarias):
    """Consolida a execução orçamentária e cruza com o faturamento do SAG."""
    if not orc:
        return None
    reg, tot_ug, periodos = orc["reg"], orc["total_ug"], orc["periodos"]
    anos = [a for a in orc["anos"] if periodos[a]["fechado"]]     # exercícios encerrados
    parciais = [a for a in orc["anos"] if not periodos[a]["fechado"]]

    def v(ano, met, ug=None, pi=None):
        return sum(x for (a, u, p, m), x in reg.items()
                   if a == ano and m == met and (ug is None or u == ug) and (pi is None or p == pi))

    E, L, P = "DESPESAS EMPENHADAS", "DESPESAS LIQUIDADAS", "PAGAMENTOS TOTAIS (EXERCICIO + RP)"

    # --- por serviço e ano
    servicos = []
    for pi, rot, chave in PIS:
        linha = {"pi": pi, "rotulo": rot, "chave": chave, "anos": {}}
        for a in anos:
            linha["anos"][a] = {"emp": v(a, E, pi=pi), "liq": v(a, L, pi=pi), "pago": v(a, P, pi=pi)}
        linha["ativo"] = any(x["emp"] or x["pago"] for x in linha["anos"].values())
        linha["var_pago"] = var(linha["anos"][anos[0]]["pago"], linha["anos"][anos[-1]]["pago"])
        servicos.append(linha)

    totais = {a: {"emp": v(a, E), "liq": v(a, L), "pago": v(a, P)} for a in anos}

    # --- cadeia de restos a pagar: o que cada exercício empurrou para o seguinte
    #     (conferido: empenhado − liquidado do ano N = RAP-NP inscrito lançado em N+1)
    rap = []
    for i, a in enumerate(anos):
        levado = totais[a]["emp"] - totais[a]["liq"]
        recebido = v(a, "RESTOS A PAGAR NAO PROCESSADOS INSCRITOS")
        confere = (abs(levado - totais[anos[i - 1]]["emp"] + totais[anos[i - 1]]["liq"]) < 0.01
                   if i else None)
        rap.append({"ano": a, "levado": levado, "recebido_do_anterior": recebido,
                    "liq_sobre_emp": (totais[a]["liq"] / totais[a]["emp"] * 100)
                                     if totais[a]["emp"] else None, "confere": confere})

    # --- peso das concessionárias no custeio de cada UG
    peso = []
    for u in UGS:
        item = {"ug": u["ug"], "sigla": u["sigla"], "logo": u["logo"], "anos": {}}
        for a in anos:
            c = v(a, L, ug=u["ug"])
            t = tot_ug.get((a, u["ug"], L), 0.0)
            item["anos"][a] = {"conc": c, "total": t, "pct": (c / t * 100) if t else None}
        peso.append(item)
    peso_total = {}
    for a in anos:
        c = v(a, L)
        t = sum(tot_ug.get((a, u["ug"], L), 0.0) for u in UGS)
        peso_total[a] = {"conc": c, "total": t, "pct": (c / t * 100) if t else None}

    # --- validação cruzada: SAG (faturado no ano) x caixa do ano
    def sis_agua(ug, a):
        return soma(series[ug]["rs_agua"][a])

    def sis_energia(ug, a):
        s = series[ug]
        if ug in tarifarias:
            return sum((s["kwh_ponta"][a][i] or 0) * (s["rs_ponta"][a][i] or 0) +
                       (s["kwh_fora"][a][i] or 0) * (s["rs_fora"][a][i] or 0) for i in range(12))
        return soma(s["rs_ponta"][a]) + soma(s["rs_fora"][a])

    anos_cruz = [a for a in anos if a in ANOS]      # só onde há série do SAG
    cruzamento = []
    for pi, rot, chave, fn in (("I3DACSPAGES", "Água e esgoto", "agua", sis_agua),
                               ("I3DACSPENEL", "Energia elétrica", "energia", sis_energia)):
        item = {"pi": pi, "rotulo": rot, "chave": chave, "anos": {}}
        for a in anos_cruz:
            s = sum(fn(u["ug"], a) for u in UGS)
            p = v(a, P, pi=pi)
            item["anos"][a] = {"sis": s, "pago": p, "razao": (p / s) if s else None}
        rs = [x["razao"] for x in item["anos"].values() if x["razao"]]
        item["razao_media"] = sum(rs) / len(rs) if rs else None
        item["por_ug"] = []
        ult = anos_cruz[-1]
        for u in UGS:
            s = fn(u["ug"], ult)
            p = v(ult, P, ug=u["ug"], pi=pi)
            item["por_ug"].append({"ug": u["ug"], "sigla": u["sigla"], "logo": u["logo"],
                                   "sis": s, "pago": p, "razao": (p / s) if s else None})
        cruzamento.append(item)

    # --- exercício em curso (posição parcial), tratado à parte dos anos fechados
    corrente = None
    if parciais:
        a = parciais[-1]
        RNP, RP_ = "RESTOS A PAGAR NAO PROCESSADOS PAGOS", "RESTOS A PAGAR PROCESSADOS PAGOS"
        INS, REI = ("RESTOS A PAGAR NAO PROCESSADOS INSCRITOS",
                    "RESTOS A PAGAR NAO PROCESSADOS REINSCRITOS")
        CAN = "RESTOS A PAGAR NAO PROCESSADOS CANCELADOS"
        exerc, rnp, rp_ = v(a, "DESPESAS PAGAS"), v(a, RNP), v(a, RP_)
        total = v(a, P)
        ins, rei, can = v(a, INS), v(a, REI), v(a, CAN)
        corrente = {
            "ano": a, "mes": periodos[a]["mes"], "rotulo": periodos[a]["rotulo"],
            "n_meses": periodos[a]["n_meses"],
            "provisao": v(a, "PROVISAO RECEBIDA"), "empenhado": v(a, E), "liquidado": v(a, L),
            "pago_total": total,
            "composicao": [
                {"rot": "Despesa do próprio exercício", "v": exerc,
                 "pct": exerc / total * 100 if total else None},
                {"rot": "Restos a pagar não processados", "v": rnp,
                 "pct": rnp / total * 100 if total else None},
                {"rot": "Restos a pagar processados", "v": rp_,
                 "pct": rp_ / total * 100 if total else None},
            ],
            "pct_herdado": (rnp + rp_) / total * 100 if total else None,
            "rap_herdado": ins + rei - can,
            "rap_inscritos": ins, "rap_reinscritos": rei, "rap_cancelados": can,
            "rap_quitado": rnp,
            "rap_saldo": ins + rei - can - rnp,
            "rap_pct_quitado": rnp / (ins + rei - can) * 100 if (ins + rei - can) else None,
            "confere": abs(exerc + rnp + rp_ - total) < 0.01,
            "servicos": [],
        }
        for pi, rot, chave in PIS:
            e, l = v(a, E, pi=pi), v(a, L, pi=pi)
            item = {"pi": pi, "rotulo": rot, "chave": chave, "emp": e, "liq": l,
                    "rap_pago": v(a, RNP, pi=pi), "pago": v(a, P, pi=pi),
                    "rap_herdado": v(a, INS, pi=pi)}
            if item["emp"] or item["pago"] or item["rap_herdado"]:
                corrente["servicos"].append(item)

    return {"anos": anos, "periodos": periodos, "anos_cruzamento": anos_cruz,
            "servicos": servicos, "totais": totais, "rap": rap, "peso": peso,
            "peso_total": peso_total, "cruzamento": cruzamento, "corrente": corrente,
            "caixa_total": {a: sum(x["anos"][a]["pago"] for x in servicos) for a in anos}}


def completude(series):
    rotulos = {"kwh_ponta": "Energia — kWh ponta", "kwh_fora": "Energia — kWh fora ponta",
               "rs_ponta": "Energia — R$ ponta", "rs_fora": "Energia — R$ fora ponta",
               "m3": "Água — m³", "rs_agua": "Água — R$ fatura"}
    out = []
    for u in UGS:
        s = series[u["ug"]]
        for k, rot in rotulos.items():
            linha = {"ug": u["ug"], "sigla": u["sigla"], "metrica": rot, "anos": {}}
            for a in ANOS:
                m = meses_com_dado(s[k][a])
                linha["anos"][a] = {"n": len(m), "faltando": [x for x in MES if x not in m]}
            out.append(linha)
    return out


# ------------------------------------------------------------------------ achados

def achados(series, tarifarias, dec, ms, ser_agua, orc):
    s238, s246 = series["160238"], series["160246"]
    tar_ponta_238 = s238["rs_fora"]["2026"][0] / s238["kwh_fora"]["2026"][0]
    tar_fora_238 = s238["rs_ponta"]["2026"][0] / s238["kwh_ponta"]["2026"][0]
    tar_ponta_246 = s246["rs_ponta"]["2026"][0] / s246["kwh_ponta"]["2026"][0]
    kwh_ponta_decl = soma(s238["kwh_ponta"]["2026"], 6)
    custo_ponta = kwh_ponta_decl * tar_ponta_238
    custo_fora = kwh_ponta_decl * tar_fora_238
    m238 = ms["energia"][0]

    a = []
    a.append({
        "id": "A1", "grau": "crítico", "servico": "Energia", "ug": "160238",
        "sigla": "Ba Ap Log Ex",
        "titulo": "Colunas PONTA e FORA PONTA invertidas em 2026",
        "resumo": "Todo o lançamento de energia de 2026 da Ba Ap Log Ex está com as colunas de "
                  "ponta e fora ponta trocadas — nos kWh e também nos reais. Como a troca é "
                  "simétrica, o total do mês não é afetado; a abertura por posto tarifário, sim, "
                  "e com ela todo indicador que dependa dessa separação.",
        "evidencias": [
            f"Jan/2026 registra {fmt(s238['kwh_ponta']['2026'][0])} kWh na ponta contra "
            f"{fmt(s238['kwh_ponta']['2025'][0])} kWh em jan/2025 — salto de "
            f"{fmt(s238['kwh_ponta']['2026'][0] / s238['kwh_ponta']['2025'][0], 1)}× sem nenhum "
            f"fato gerador correspondente.",
            f"A coluna fora ponta de jan/2026 traz {fmt(s238['kwh_fora']['2026'][0])} kWh, "
            f"praticamente idêntico à ponta de jan/2025 "
            f"({fmt(s238['kwh_ponta']['2025'][0])} kWh — diferença de "
            f"{fmt(abs(var(s238['kwh_ponta']['2025'][0], s238['kwh_fora']['2026'][0])), 1)}%). "
            f"As duas colunas simplesmente trocaram de lugar.",
            f"A tarifa implícita fecha a prova: a coluna “fora ponta” de jan/2026 resulta em "
            f"R$ {fmt(tar_ponta_238, 4)}/kWh, patamar de ponta — a mesma tarifa de ponta apurada "
            f"no DC Mun no mesmo mês (R$ {fmt(tar_ponta_246, 4)}/kWh). Duas UG da mesma praça e "
            f"do mesmo mês: o que está trocado é a coluna.",
        ],
        "impacto": f"O acumulado de ponta da UG aparece em {fmt(kwh_ponta_decl)} kWh contra uma "
                   f"meta de {fmt(m238['ponta_meta'])} kWh — desvio aparente de "
                   f"{pc(m238['ponta_desvio'], 0)}, que não existe. E, se a classificação chegar "
                   f"assim à fatura, esses {fmt(kwh_ponta_decl)} kWh custariam "
                   f"R$ {fmt(custo_ponta, 2)} no posto de ponta contra R$ {fmt(custo_fora, 2)} "
                   f"fora dele — R$ {fmt(custo_ponta - custo_fora, 2)} em seis meses.",
        "acao": "Reprocessar os seis meses de 2026 no SAG e conferir contra as faturas "
                "originais da distribuidora (demanda contratada e postos tarifários).",
    })

    sig = " e ".join(UGIDX[u]["sigla"] for u in tarifarias)
    ex = series[tarifarias[0]]
    ex_sig = UGIDX[tarifarias[0]]["sigla"]
    ex_rs = soma(ex["rs_ponta"]["2026"], 6) + soma(ex["rs_fora"]["2026"], 6)
    ex_kwh = soma(ex["kwh_ponta"]["2026"], 6) + soma(ex["kwh_fora"]["2026"], 6)
    ex_rec = sum((ex["kwh_ponta"]["2026"][i] or 0) * (ex["rs_ponta"]["2026"][i] or 0) +
                 (ex["kwh_fora"]["2026"][i] or 0) * (ex["rs_fora"]["2026"][i] or 0)
                 for i in range(6))
    a.append({
        "id": "A2", "grau": "crítico", "servico": "Energia", "ug": "/".join(tarifarias),
        "sigla": sig,
        "titulo": "Campo VALOR preenchido com a TARIFA (R$/kWh), não com o valor da fatura",
        "resumo": f"{sig} lançam no campo de valor a tarifa unitária. O relatório consolidado "
                  f"soma isso como se fosse dinheiro, e o gasto dessas duas UG praticamente "
                  f"desaparece do total do conjunto.",
        "evidencias": [
            f"{ex_sig}: R$ {fmt(ex_rs, 2)} de “gasto” em jan–jun/2026 para {fmt(ex_kwh)} kWh "
            f"consumidos — o equivalente a R$ {fmt(ex_rs / (ex_kwh / 1000), 2)} por MWh.",
            f"Os números mensais lançados como valor de ponta em 2026 "
            f"({', '.join(fmt(v, 2) for v in ex['rs_ponta']['2026'][:6])}) têm ordem de grandeza "
            f"de tarifa, não de fatura.",
            f"Confirmação cruzada: o “valor” de ponta de jan/2026 dessas UG "
            f"(R$ {fmt(ex['rs_ponta']['2026'][0], 2)}) coincide com a tarifa de ponta apurada nas "
            f"demais UG no mesmo mês (R$ {fmt(tar_ponta_246, 4)}/kWh).",
        ],
        "impacto": f"Reconstruído por kWh × tarifa, só o 1º semestre de 2026 do {ex_sig} soma "
                   f"R$ {fmt(ex_rec, 2)} — valor que hoje não aparece em lugar nenhum do "
                   f"consolidado.",
        "acao": "Padronizar o preenchimento do SAG: campo VALOR = valor total da fatura, "
                "com tributos, encargos e bandeira tarifária. Recompor a série 2023–2026 dessas "
                "duas UG.",
    })

    s321 = series["160321"]
    m321 = [m for m in ms["energia"] if m["ug"] == "160321"][0]
    a.append({
        "id": "A3", "grau": "alto", "servico": "Energia", "ug": "160321", "sigla": "ECT",
        "titulo": "Série do ECT muda de posto tarifário entre 2024 e 2025",
        "resumo": "Até 2024 todo o consumo do ECT foi lançado como ponta; a partir de 2025 "
                  "passou a ser lançado integralmente como fora ponta. A série não é comparável "
                  "entre 2024 e 2025, e a meta de 2026 herdou o critério antigo.",
        "evidencias": [
            f"2024: {fmt(soma(s321['kwh_ponta']['2024']))} kWh na ponta e "
            f"{fmt(soma(s321['kwh_fora']['2024']))} kWh fora ponta.",
            f"2025: {fmt(soma(s321['kwh_ponta']['2025']))} kWh na ponta e "
            f"{fmt(soma(s321['kwh_fora']['2025']))} kWh fora ponta.",
            "Nov e dez/2024 estão sem valor lançado (R$ 0,00) com consumo registrado.",
        ],
        "impacto": f"A meta de 2026 do ECT ainda reserva {fmt(m321['ponta_meta'])} kWh para a "
                   f"ponta no período apurado, onde hoje não se lança nada — por isso o ECT "
                   f"aparece {pc(m321['desvio'])} em relação à meta total, número sem sentido. "
                   f"Comparado só no fora ponta, o ECT está {pc(m321['fora_desvio'])} "
                   f"em relação à sua meta. E essa meta de fora ponta é frágil: como 2023 e 2024 "
                   f"estão zerados, o SAG a calcula sobre <b>um único ano</b> (2025), sem a "
                   f"suavização que a média de três anos daria.",
        "acao": "Confirmar junto ao ECT se houve mudança de modalidade tarifária ou apenas erro "
                "de coluna, refazer 2023–2024 no critério atual e recalcular a meta.",
    })

    s304 = series["160304"]
    m304 = [m for m in ms["agua"] if m["ug"] == "160304"][0]
    tar_abr = s304["rs_fora"]["2025"][3] / s304["kwh_fora"]["2025"][3]
    tar_24 = soma(s304["rs_fora"]["2024"]) / soma(s304["kwh_fora"]["2024"])
    med_mes_24 = soma(s304["kwh_fora"]["2024"]) / 12
    a.append({
        "id": "A4", "grau": "alto", "servico": "Energia + Água", "ug": "160304", "sigla": "BMSA",
        "titulo": "Refaturamentos e fatura não lançada contaminam a base de 2025",
        "resumo": "A série de 2025 do BMSA tem meses com consumo e valor incompatíveis entre si, "
                  "típicos de refaturamento retroativo, e um mês de água sem lançamento. Isso "
                  "contamina a média 2023–2025 que gera a meta de 2026.",
        "evidencias": [
            f"Abr/2025: {fmt(s304['kwh_fora']['2025'][3])} kWh fora ponta com "
            f"R$ {fmt(s304['rs_fora']['2025'][3], 2)} — tarifa implícita de "
            f"R$ {fmt(tar_abr, 2)}/kWh, {fmt(tar_abr / tar_24, 1)}× a tarifa média de 2024.",
            f"Nov/2025: {fmt(s304['kwh_fora']['2025'][10])} kWh fora ponta "
            f"({fmt(s304['kwh_fora']['2025'][10] / med_mes_24, 1)}× a média mensal de 2024) e "
            f"dez/2025: {fmt(s304['kwh_fora']['2025'][11])} kWh "
            f"({fmt(s304['kwh_fora']['2025'][11] / med_mes_24, 1)}×).",
            "Nov/2025 da água está zerado em consumo e em valor: fatura não lançada.",
        ],
        "impacto": f"O BMSA aparece {pc(m304['desvio'])} em relação à meta de água — o maior "
                   f"desvio do conjunto — sobre uma base de comparação com um mês faltando e "
                   f"outros inflados por refaturamento.",
        "acao": "Marcar os meses de refaturamento no SAG, lançar nov/2025 da água e "
                "recalcular a média trienal excluindo os atípicos.",
    })

    a.append({
        "id": "A5", "grau": "médio", "servico": "Energia + Água", "ug": "TODAS",
        "sigla": "Conjunto",
        "titulo": "Meses de fechamento diferentes entre as UG impedem consolidação direta",
        "resumo": "O relatório consolidado soma UG fechadas em meses diferentes, o que produz um "
                  "acumulado do conjunto sem significado.",
        "evidencias": [
            "Energia: BMSA e ECT param em mai/2026; Ba Ap Log Ex, DC Mun e 1º D Sup vão até "
            "jun/2026; o BCMS já tem jul/2026 lançado.",
            "Água: cinco UG até jun/2026 e o BCMS até jul/2026.",
            "No relatório consolidado de energia há 35 registros mensais, sendo 34 no intervalo "
            "jan–jun: faltam 2 dos 36 esperados (6 UG × 6 meses).",
        ],
        "impacto": "Toda comparação do conjunto neste relatório foi restrita a jan–jun e, na "
                   "energia, a jan–mai onde a UG não tem junho lançado. A base usada está "
                   "indicada em cada tabela.",
        "acao": "Fixar o mês de corte do relatório e só consolidar UG fechadas até ele.",
    })

    piores = sorted([i for i in dec["itens"] if i["var_p"] and i["var_p"] > 0],
                    key=lambda x: -x["var_p"])
    a.append({
        "id": "A6", "grau": "informativo", "servico": "Água", "ug": "TODAS", "sigla": "Conjunto",
        "titulo": "A conta de água sobe por preço, não por consumo",
        "resumo": "Este é o único bloco em que os dados sustentam, sem ressalva, a tese de que se "
                  "está pagando mais caro.",
        "evidencias": [
            f"Tarifa média efetiva do conjunto: R$ {fmt(dec['total']['p25'], 2)}/m³ em 2025 para "
            f"R$ {fmt(dec['total']['p26'], 2)}/m³ em 2026 ({pc(dec['total']['var_p'])}). Desde "
            f"2023 (R$ {fmt(ser_agua[0]['tarifa'], 2)}/m³) a alta acumulada é de "
            f"{pc(var(ser_agua[0]['tarifa'], dec['total']['p26']))}, em valores nominais.",
            f"{len(piores)} das 6 UG tiveram alta de tarifa, com destaque para "
            f"{piores[0]['sigla']} ({pc(piores[0]['var_p'])}).",
            f"Efeito preço no semestre: R$ {fmt(dec['total']['efeito_preco'], 2)}. Mantido o "
            f"mesmo comportamento no 2º semestre, seriam cerca de "
            f"R$ {fmt(dec['total']['efeito_preco'] * 2, 0)} no ano.",
        ],
        "impacto": f"O conjunto consumiu {fmt(abs(dec['total']['var_m']), 1)}% menos água e ainda "
                   f"assim só reduziu {fmt(abs(dec['total']['var_r']), 1)}% da despesa. O aumento "
                   f"de preço anulou "
                   f"{fmt(abs(dec['total']['efeito_preco'] / dec['total']['efeito_volume']) * 100)}% "
                   f"da economia obtida com a redução de consumo.",
        "acao": "Não há ação corretiva de dado. As alavancas disponíveis são consumo e revisão "
                "contratual/cadastral junto à concessionária.",
    })

    if orc:
        ce, ca = orc["cruzamento"][1], orc["cruzamento"][0]
        anos_c = orc["anos_cruzamento"]
        razoes = ", ".join(f"{fmt(ce['anos'][x]['razao'], 3)} em {x}" for x in anos_c)
        razoes_a = ", ".join(f"{fmt(ca['anos'][x]['razao'], 3)} em {x}" for x in anos_c)
        disp = sorted(ce["por_ug"], key=lambda x: -(x["razao"] or 0))
        a.append({
            "id": "A7", "grau": "alto", "servico": "Energia", "ug": "TODAS", "sigla": "Conjunto",
            "titulo": "O SAG registra cerca de 84% do que a fatura de energia custa",
            "resumo": "Confrontado com o caixa efetivamente pago, o faturamento de energia "
                      "lançado no SAG fica sistematicamente abaixo do desembolso. Na água "
                      "isso não acontece — lá o sistema bate com o caixa.",
            "evidencias": [
                f"Razão entre o pago no ano e o faturado no SAG, energia: {razoes}. "
                f"A estabilidade em três exercícios indica componente estrutural, não erro "
                f"pontual.",
                f"Na água a mesma razão é {razoes_a} — praticamente 1 para 1. É o que valida o "
                f"dado de água e isola o problema na energia.",
                f"A dispersão por UG é grande em {anos_c[-1]}: de "
                f"{fmt(disp[-1]['razao'], 2)} no {disp[-1]['sigla']} a "
                f"{fmt(disp[0]['razao'], 2)} na {disp[0]['sigla']}, o que sugere dependência da "
                f"modalidade tarifária de cada unidade.",
            ],
            "impacto": "O SAG mede consumo faturado em kWh. A fatura cobra também demanda "
                       "contratada, energia reativa excedente, bandeiras e a contribuição de "
                       "iluminação pública — nenhum desses itens entra no sistema. Planejar "
                       "energia só pelo SAG subestima a despesa em torno de um sexto.",
            "acao": "Confrontar uma fatura completa de cada UG com o respectivo lançamento para "
                    "quantificar cada componente ausente. Onde houver demanda contratada acima "
                    "do uso efetivo, há economia disponível sem reduzir consumo.",
        })

        ult, ant = orc["rap"][-1], orc["rap"][-2]
        cr = orc["corrente"]
        ev = [
            f"{ult['ano']}: empenhado R$ {fmt(orc['totais'][ult['ano']]['emp'], 2)}, "
            f"liquidado R$ {fmt(orc['totais'][ult['ano']]['liq'], 2)} "
            f"({fmt(ult['liq_sobre_emp'], 1)}%) — restam R$ {fmt(ult['levado'], 2)}.",
            f"É {fmt(ult['levado'] / ant['levado'], 1)}× o que {ant['ano']} levou para "
            f"{int(ant['ano']) + 1} (R$ {fmt(ant['levado'], 2)}).",
            "A relação foi conferida ano a ano: o empenhado menos o liquidado de cada exercício "
            "reaparece exatamente como resto a pagar não processado inscrito no seguinte.",
        ]
        impacto = ("O saldo não consome crédito novo, mas consome limite financeiro. Somado ao "
                   "aumento real de tarifa, ajuda a explicar por que o exercício corrente aperta "
                   "mesmo com consumo estável ou em queda.")
        if cr and cr["ano"] == str(int(ult["ano"]) + 1):
            agua = next((s for s in cr["servicos"] if s["chave"] == "agua"), None)
            ev.append(
                f"O crosstab de {cr['rotulo']} confirma o valor <b>ao centavo</b>: os restos a "
                f"pagar não processados inscritos em {cr['ano']} somam exatamente "
                f"R$ {fmt(ult['levado'], 2)}.")
            if agua and not agua["emp"] and not agua["liq"]:
                ev.append(
                    f"Em {cr['rotulo']}, a água das seis UG roda <b>inteiramente sobre o resto a "
                    f"pagar de {ult['ano']}</b>: R$ 0,00 empenhados e R$ 0,00 liquidados no "
                    f"exercício, com R$ {fmt(agua['rap_pago'], 2)} pagos contra o saldo herdado.")
            impacto = (f"Em {cr['rotulo']}, <b>{fmt(cr['pct_herdado'], 1)}% de tudo que as seis UG "
                       f"desembolsaram com concessionárias</b> foi para quitar exercícios "
                       f"anteriores — R$ {fmt(cr['pago_total'] * cr['pct_herdado'] / 100, 2)} de "
                       f"um total de R$ {fmt(cr['pago_total'], 2)}. A despesa própria de "
                       f"{cr['ano']} mal começou a ser liquidada "
                       f"(R$ {fmt(cr['liquidado'], 2)}), o que significa que o custo real do "
                       f"exercício ainda está por aparecer.")
        a.append({
            "id": "A8", "grau": "crítico", "servico": "Orçamento", "ug": "TODAS",
            "sigla": "Conjunto",
            "titulo": f"{ult['ano']} empurrou R$ {fmt(ult['levado'], 2)} de concessionárias "
                      f"para {int(ult['ano']) + 1}",
            "resumo": f"O empenho de {ult['ano']} foi muito além do que se conseguiu liquidar no "
                      f"exercício. A diferença virou resto a pagar e passou a pressionar o "
                      f"financeiro do ano seguinte — parte da conta que {int(ult['ano']) + 1} "
                      f"parece estar pagando é, na verdade, de {ult['ano']}.",
            "evidencias": ev,
            "impacto": impacto,
            "acao": "Dimensionar o empenho estimativo pela liquidação histórica, não pelo teto "
                    "do contrato, e acompanhar o estoque de restos a pagar por concessionária "
                    "junto com o consumo.",
        })
    return a


# ------------------------------------------------------------------------- HTML

def gerar_html(d):
    return TEMPLATE.replace("__DADOS__", json.dumps(d, ensure_ascii=False))


TEMPLATE = r"""<!doctype html>
<html lang="pt-BR">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Concessionárias 2026 — BA Ap Log Ex e OMDS | Relatório Consolidado</title>
<meta name="description" content="Histórico de consumo, tarifas e despesa com concessionárias das seis UG apoiadas pela BA Ap Log Ex — 1ª RM. Base: SAG e Tesouro Gerencial.">
<link rel="icon" href="assets/e10-mini.jpg">
<meta name="theme-color" content="#132840">
<style>
/* Tema derivado do emblema do E10 — Gerenciamento de Dados:
   azul-marinho profundo do fundo, ouro do escudo e ciano dos elementos de dados. */
:root{
  --verde:#132840; --verde-2:#1b4f72; --oliva:#7a5a12; --areia:#c9971b;
  --tinta:#0f1826; --papel:#f5f4ef; --cartao:#fff; --linha:#e0dccf;
  --texto:#2a3340; --suave:#616c7c;
  --alta:#b3261e; --media:#96660a; --baixa:#15705a; --info:#0e6f96;
  --c1:#1b4f72; --c2:#0e7c9e; --c3:#a97c09; --c4:#c1440e; --c5:#5d6d7e; --c6:#8a6d3b;
  --sobe:#b3261e; --desce:#15705a;
  --sombra:0 1px 2px rgba(15,24,38,.07),0 8px 24px rgba(15,24,38,.07);
}
@media (prefers-color-scheme:dark){
  :root{--papel:#0c1420;--cartao:#141f30;--linha:#26364b;--tinta:#f2f4f7;
        --texto:#dae1ea;--suave:#95a2b3;--verde:#8ec5f0;--verde-2:#a9d6f7;--areia:#e5b219;
        --oliva:#e5b219;
        --alta:#ff7a6e;--media:#e8a33f;--baixa:#4fd1a0;--info:#5dade2;
        --c1:#5dade2;--c2:#48c9d6;--c3:#e5b219;--c4:#e8834a;--c5:#a6b5c4;--c6:#c9a227;
        --sobe:#ff7a6e;--desce:#4fd1a0;
        --sombra:0 1px 2px rgba(0,0,0,.5),0 8px 24px rgba(0,0,0,.4);}
}
*{box-sizing:border-box}
body{margin:0;background:var(--papel);color:var(--texto);
  font:16px/1.6 -apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,"Helvetica Neue",Arial,sans-serif;
  -webkit-font-smoothing:antialiased}
.wrap{max-width:1120px;margin:0 auto;padding:0 20px}
a{color:var(--info)}
h1,h2,h3{color:var(--tinta);line-height:1.2;margin:0 0 .4em}
h1{font-size:clamp(26px,4.2vw,44px);letter-spacing:-.02em}
h2{font-size:clamp(20px,2.6vw,29px);letter-spacing:-.01em;margin-top:0}
h3{font-size:17px}
p{margin:0 0 1em}
.lead{font-size:clamp(16px,1.9vw,19px);opacity:.92}
.sub{color:var(--suave);font-size:14px}

header.topo{background:linear-gradient(155deg,#16304d 0%,var(--verde) 45%,#08111d 100%);
  color:#eef2f7;padding:44px 0 38px;border-bottom:4px solid var(--areia);position:relative}
header.topo::after{content:"";position:absolute;left:0;right:0;bottom:-4px;height:4px;
  background:linear-gradient(90deg,#8a6a10,#e5b219 18%,#fff0b8 50%,#e5b219 82%,#8a6a10)}
@media (prefers-color-scheme:dark){
  header.topo{background:linear-gradient(155deg,#122540,#0a1524 55%,#060d16)}}
header.topo h1{color:#fff;margin-bottom:10px}
header.topo .lead{color:#cfdae7;max-width:64ch}
.cabeca{display:flex;gap:30px;align-items:flex-start}
.cabeca .texto{flex:1;min-width:0}
.emblema{flex:none;width:168px;height:auto;border-radius:14px;display:block;
  border:2px solid rgba(229,178,25,.55);box-shadow:0 6px 26px rgba(0,0,0,.42)}
@media (max-width:760px){
  .cabeca{flex-direction:column-reverse;gap:20px}
  .emblema{width:132px;align-self:flex-start}
}
.crumb{font:600 12px/1 ui-monospace,SFMono-Regular,Menlo,monospace;letter-spacing:.14em;
  text-transform:uppercase;color:var(--areia);margin-bottom:14px}
.marca{display:flex;gap:8px;flex-wrap:wrap;margin-top:22px}
.marca img{height:40px;width:40px;object-fit:contain;background:#fff;border-radius:8px;padding:4px}

nav.indice{position:sticky;top:0;z-index:20;background:var(--cartao);border-bottom:1px solid var(--linha);
  overflow-x:auto;-webkit-overflow-scrolling:touch}
nav.indice ul{display:flex;gap:2px;list-style:none;margin:0 auto;padding:0 20px;max-width:1120px;
  white-space:nowrap}
nav.indice a{display:block;padding:13px 12px;font-size:13.5px;font-weight:600;color:var(--suave);
  text-decoration:none;border-bottom:2px solid transparent}
nav.indice a:hover{color:var(--tinta)}
nav.indice a.on{color:var(--verde-2);border-bottom-color:var(--verde-2)}

section{padding:52px 0;border-bottom:1px solid var(--linha)}
section:last-of-type{border-bottom:0}
.tag{display:inline-block;font:700 11px/1 ui-monospace,monospace;letter-spacing:.12em;
  text-transform:uppercase;color:var(--oliva);margin-bottom:10px}
@media (prefers-color-scheme:dark){.tag{color:var(--areia)}}

.kpis{display:grid;grid-template-columns:repeat(auto-fit,minmax(190px,1fr));gap:14px;margin:26px 0}
.kpi{background:var(--cartao);border:1px solid var(--linha);border-radius:12px;padding:16px 18px;
  box-shadow:var(--sombra)}
.kpi .rot{font-size:12.5px;color:var(--suave);text-transform:uppercase;letter-spacing:.05em;
  font-weight:700;margin-bottom:6px}
.kpi .num{font:700 27px/1.1 ui-monospace,SFMono-Regular,Menlo,monospace;color:var(--tinta);
  letter-spacing:-.02em}
.kpi .pe{font-size:13px;color:var(--suave);margin-top:6px}
.pos,.kpi .num.pos,.mini .v.pos{color:var(--alta)}
.neg,.kpi .num.neg,.mini .v.neg{color:var(--baixa)}

.cartao{background:var(--cartao);border:1px solid var(--linha);border-radius:14px;
  padding:22px;box-shadow:var(--sombra);margin:18px 0}
.aviso{border-left:4px solid var(--media);background:color-mix(in srgb,var(--media) 7%,var(--cartao))}
.perigo{border-left:4px solid var(--alta);background:color-mix(in srgb,var(--alta) 7%,var(--cartao))}
.ok{border-left:4px solid var(--baixa);background:color-mix(in srgb,var(--baixa) 7%,var(--cartao))}

table{width:100%;border-collapse:collapse;font-size:14px}
.rolagem{overflow-x:auto;-webkit-overflow-scrolling:touch;margin:14px 0;
  border:1px solid var(--linha);border-radius:12px;background:var(--cartao)}
.rolagem table{min-width:640px}
th,td{padding:10px 12px;text-align:right;border-bottom:1px solid var(--linha);white-space:nowrap}
th:first-child,td:first-child{text-align:left;white-space:normal}
thead th{background:color-mix(in srgb,var(--verde) 8%,var(--cartao));color:var(--tinta);
  font-size:12px;text-transform:uppercase;letter-spacing:.05em;position:sticky;top:0}
tbody tr:last-child td{border-bottom:0}
tfoot td{font-weight:700;background:color-mix(in srgb,var(--verde) 5%,var(--cartao));
  border-top:2px solid var(--linha)}
td.num,th.num{font-family:ui-monospace,SFMono-Regular,Menlo,monospace}
.uni{display:flex;align-items:center;gap:9px}
.uni img{height:26px;width:26px;object-fit:contain;background:#fff;border-radius:6px;padding:2px;flex:none}

.pilula{display:inline-block;padding:2px 9px;border-radius:999px;font:700 11.5px/1.7 sans-serif;
  text-transform:uppercase;letter-spacing:.04em}
.p-alta{background:color-mix(in srgb,var(--alta) 15%,transparent);color:var(--alta)}
.p-media{background:color-mix(in srgb,var(--media) 15%,transparent);color:var(--media)}
.p-baixa{background:color-mix(in srgb,var(--baixa) 15%,transparent);color:var(--baixa)}
.p-info{background:color-mix(in srgb,var(--info) 15%,transparent);color:var(--info)}

.achado{border:1px solid var(--linha);border-radius:14px;background:var(--cartao);margin:14px 0;
  overflow:hidden;box-shadow:var(--sombra)}
.achado > summary{padding:16px 20px;cursor:pointer;list-style:none;display:flex;gap:12px;
  align-items:flex-start;font-weight:650}
.achado > summary::-webkit-details-marker{display:none}
.achado > summary::after{content:"+";margin-left:auto;font:700 20px/1 monospace;color:var(--suave)}
.achado[open] > summary::after{content:"\2212"}
.achado .corpo{padding:0 20px 20px;border-top:1px solid var(--linha);padding-top:16px}
.achado ul{margin:.4em 0 1em;padding-left:1.2em}
.achado li{margin-bottom:.45em}
.rot-mini{font:700 11px/1 ui-monospace,monospace;letter-spacing:.1em;text-transform:uppercase;
  color:var(--suave);display:block;margin:14px 0 5px}

.grade{display:grid;grid-template-columns:repeat(auto-fit,minmax(320px,1fr));gap:16px;margin-top:20px}
.card-ug{background:var(--cartao);border:1px solid var(--linha);border-radius:14px;padding:18px;
  box-shadow:var(--sombra)}
.card-ug header{display:flex;gap:11px;align-items:center;margin-bottom:14px}
.card-ug header img{height:38px;width:38px;object-fit:contain;background:#fff;border-radius:8px;padding:3px}
.card-ug h3{margin:0;font-size:16px}
.card-ug .nomeug{font-size:13px;line-height:1.3;color:var(--texto);margin:1px 0 2px}
.card-ug .ugid{font:12px/1.3 ui-monospace,monospace;color:var(--suave)}
.mini{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-top:12px}
.mini div{background:color-mix(in srgb,var(--verde) 5%,transparent);border-radius:9px;padding:9px 11px}
.mini .r{font-size:11.5px;color:var(--suave);text-transform:uppercase;letter-spacing:.04em;font-weight:700}
.mini .v{font:700 17px/1.2 ui-monospace,monospace;color:var(--tinta)}

svg .eixo{stroke:var(--linha);stroke-width:1}
svg .rot{fill:var(--suave);font:11px ui-monospace,monospace}
svg .rotb{fill:var(--tinta);font:700 11px ui-monospace,monospace}
svg .grade-h{stroke:var(--linha);stroke-width:1;stroke-dasharray:3 4}
.legenda{text-align:center;margin:-4px 0 0}

.matriz{font-size:12.5px}
.matriz td{padding:6px 8px}
.cel{display:inline-block;min-width:34px;text-align:center;border-radius:5px;padding:2px 5px;
  font:700 11.5px/1.5 ui-monospace,monospace}
.c-ok{background:color-mix(in srgb,var(--baixa) 16%,transparent);color:var(--baixa)}
.c-par{background:color-mix(in srgb,var(--media) 18%,transparent);color:var(--media)}
.c-nao{background:color-mix(in srgb,var(--alta) 15%,transparent);color:var(--alta)}
.c-neu{background:color-mix(in srgb,var(--suave) 14%,transparent);color:var(--suave)}

ol.passos{counter-reset:p;list-style:none;padding:0;margin:18px 0}
ol.passos li{counter-increment:p;position:relative;padding:14px 0 14px 52px;border-bottom:1px solid var(--linha)}
ol.passos li:last-child{border-bottom:0}
ol.passos li::before{content:counter(p);position:absolute;left:0;top:13px;width:32px;height:32px;
  border-radius:50%;background:var(--verde);color:#fff;display:grid;place-items:center;
  font:700 14px/1 sans-serif}
@media (prefers-color-scheme:dark){ol.passos li::before{background:var(--verde-2);color:#12161a}}
ol.passos b{color:var(--tinta)}

footer{background:var(--verde);color:#dfe6dd;padding:34px 0;font-size:13.5px}
@media (prefers-color-scheme:dark){footer{background:#16281d}}
footer a{color:var(--areia)}
footer .creditos{margin:0 0 18px;padding:16px 0;font-size:14.5px;line-height:1.5;
  border-top:1px solid rgba(229,178,25,.32);border-bottom:1px solid rgba(229,178,25,.32);
  display:flex;gap:16px;align-items:center}
footer .creditos img{flex:none;width:62px;height:auto;border-radius:9px;
  border:1px solid rgba(229,178,25,.5)}
footer .creditos b{color:var(--areia)}
@media (max-width:520px){footer .creditos{flex-direction:column;align-items:flex-start}}
@media print{
  nav.indice{display:none} section{page-break-inside:avoid;padding:24px 0}
  .achado{page-break-inside:avoid} .achado .corpo{display:block!important}
  header.topo{background:#132840!important;-webkit-print-color-adjust:exact;print-color-adjust:exact}
  .emblema{width:120px!important}
}
</style>
</head>
<body>

<header class="topo">
  <div class="wrap cabeca">
    <div class="texto">
    <div class="crumb">BA Ap Log Ex &middot; 1ª Região Militar &middot; CML &middot; 1º CGCFEx &middot; RJ</div>
    <h1>Por que a despesa com concessionárias está mais cara em 2026</h1>
    <p class="lead">Painel de acompanhamento das contas de concessionárias das seis Unidades
    Gestoras apoiadas: <b>histórico de consumo, de tarifas e de despesa</b> com água e esgoto,
    energia elétrica, telefonia e serviços postais. Não trata só do que foi pago — acompanha
    quanto se consome, a que preço e como isso evolui ao longo dos anos.</p>
    <p class="lead">São <b>duas bases de dados</b>: o <b>SAG</b> — Sistema de Acompanhamento da
    Gestão, de onde vêm o consumo medido, o faturamento e as metas (2023 a 2026); e o
    <b>Tesouro Gerencial</b>, de onde vem a execução orçamentária (2022 a 2026). Cruzar as duas é
    o que permite separar o que a unidade consumiu do que ela efetivamente pagou — e é também o
    que revela os <b>defeitos de lançamento</b> documentados aqui.</p>
    <div class="marca" id="marcas"></div>
    </div>
    <img class="emblema" src="assets/e10.jpg" width="810" height="900"
         alt="Emblema do E10 do Comando da BA Ap Log Ex — Gerenciamento de Dados">
  </div>
</header>

<nav class="indice"><ul>
  <li><a href="#resumo">Resumo</a></li>
  <li><a href="#agua">Água</a></li>
  <li><a href="#energia">Energia</a></li>
  <li><a href="#orcamento">Orçamento</a></li>
  <li><a href="#achados">Achados</a></li>
  <li><a href="#unidades">Unidades</a></li>
  <li><a href="#metas">Metas 2026</a></li>
  <li><a href="#acervo">Acervo e lacunas</a></li>
  <li><a href="#acoes">Plano de ação</a></li>
  <li><a href="#metodo">Metodologia</a></li>
</ul></nav>

<section id="resumo"><div class="wrap">
  <span class="tag">Resumo executivo</span>
  <h2>A resposta em um minuto</h2>
  <div id="resumo-txt"></div>
  <div class="kpis" id="kpis"></div>
  <div class="cartao perigo">
    <h3 style="margin-top:0">Ressalva que precede qualquer número financeiro</h3>
    <p style="margin-bottom:0" id="ressalva"></p>
  </div>
</div></section>

<section id="agua"><div class="wrap">
  <span class="tag">Bloco 1 &middot; Água e esgoto</span>
  <h2>A conta subiu porque o preço subiu</h2>
  <p class="lead" id="agua-lead"></p>
  <div id="g-tarifa"></div>
  <p class="sub legenda">Tarifa média efetiva (valor da fatura &divide; m³ medidos), jan–jun de cada
  ano. O rótulo inferior traz o volume consumido no período. Valores nominais, sem correção
  pela inflação.</p>
  <h3 style="margin-top:34px">Decomposição da variação 2025 &rarr; 2026 (jan–jun)</h3>
  <p class="sub">Quanto da variação da despesa veio do <b>preço</b> (tarifa por m³) e quanto veio do
  <b>volume</b> (m³ consumidos). Efeito preço = (tarifa 2026 &minus; tarifa 2025) &times; volume 2025.</p>
  <div id="g-cascata"></div>
  <div class="rolagem" id="t-agua"></div>
  <div class="cartao aviso" id="agua-nota"></div>
</div></section>

<section id="energia"><div class="wrap">
  <span class="tag">Bloco 2 &middot; Energia elétrica</span>
  <h2>O consumo subiu; o valor lançado não fecha</h2>
  <p class="lead" id="energia-lead"></p>
  <h3>Consumo físico — o dado confiável (kWh, jan–jun)</h3>
  <div id="g-kwh"></div>
  <div class="rolagem" id="t-kwh"></div>
  <h3 style="margin-top:34px">Valor — o dado que exige reconstrução</h3>
  <p class="sub">As UG marcadas com <b>R</b> tiveram o gasto reconstruído (kWh &times; tarifa lançada),
  porque o SAG recebeu tarifa no lugar do valor da fatura. Ver achado A2.</p>
  <div class="rolagem" id="t-rs"></div>
</div></section>

<section id="orcamento"><div class="wrap">
  <span class="tag">Bloco 3 &middot; Execução orçamentária</span>
  <h2>O que foi efetivamente pago</h2>
  <p class="lead" id="orc-lead"></p>
  <div class="kpis" id="orc-kpis"></div>

  <h3>Caixa do exercício, por serviço</h3>
  <p class="sub">Pagamentos totais do ano (despesa do exercício + restos a pagar de exercícios
  anteriores) — o desembolso real com concessionárias.</p>
  <div id="g-caixa"></div>
  <div class="rolagem" id="t-orc"></div>

  <h3 style="margin-top:34px">O SAG bate com o caixa?</h3>
  <p class="sub">Razão entre o pago no ano e o faturamento lançado no SAG. Perto de 1,00, o
  sistema reflete a despesa; acima disso, a fatura cobra o que o sistema não registra.</p>
  <div class="rolagem" id="t-cruza"></div>
  <div class="cartao" id="cruza-nota"></div>

  <h3 style="margin-top:34px">Restos a pagar — o que cada exercício empurra adiante</h3>
  <div id="g-rap"></div>
  <div class="cartao perigo" id="rap-nota"></div>

  <div id="bloco-corrente"></div>

  <h3 style="margin-top:34px">Telefonia e serviços postais</h3>
  <p class="sub">Os planos internos de menor porte, incluindo o de telefonia móvel (PTELM).
  Valores <b>pagos no ano</b> (despesa do exercício + restos a pagar).</p>
  <div class="rolagem" id="t-tel"></div>
  <div class="cartao aviso" id="tel-nota"></div>

  <h3 style="margin-top:34px">Peso das concessionárias no custeio de cada UG</h3>
  <p class="sub">Quanto da despesa liquidada de cada Unidade Gestora foi para água, energia,
  telefonia e correios.</p>
  <div class="rolagem" id="t-peso"></div>
  <p class="sub" id="peso-nota"></p>
</div></section>

<section id="achados"><div class="wrap">
  <span class="tag">Bloco 4 &middot; Auditoria do dado</span>
  <h2>Achados</h2>
  <p class="lead">Seis achados, do crítico ao informativo. Cada um traz a evidência numérica
  extraída dos próprios arquivos do SAG.</p>
  <div id="lista-achados"></div>
</div></section>

<section id="unidades"><div class="wrap">
  <span class="tag">Bloco 5 &middot; Por unidade gestora</span>
  <h2>Painel das seis UG</h2>
  <p class="lead">As seis Unidades Gestoras apoiadas pela BA Ap Log Ex, todas na
  1ª Região Militar. As denominações por extenso foram informadas pela própria Base — o SAG
  exporta apenas a sigla.</p>
  <div class="grade" id="cards-ug"></div>
</div></section>

<section id="metas"><div class="wrap">
  <span class="tag">Bloco 6 &middot; Metas</span>
  <h2>Acumulado &times; meta 2026</h2>
  <p class="lead" id="metas-lead"></p>
  <h3>Água (m³)</h3>
  <div id="g-meta-agua"></div>
  <h3 style="margin-top:30px">Energia fora ponta (kWh)</h3>
  <div id="g-meta-fora"></div>
  <h3 style="margin-top:30px">Energia ponta (kWh)</h3>
  <div id="g-meta-ponta"></div>
  <p class="sub legenda" id="ponta-legenda"></p>
  <div class="cartao aviso" id="meta-nota"></div>
  <div class="rolagem" id="t-meta"></div>
</div></section>

<section id="acervo"><div class="wrap">
  <span class="tag">Bloco 7 &middot; Acervo documental</span>
  <h2>O que existe, o que falta</h2>
  <p class="lead" id="acervo-lead"></p>
  <h3>Matriz de completude por série</h3>
  <p class="sub">Meses com lançamento em cada ano. <span class="cel c-ok">12</span> série completa &middot;
  <span class="cel c-par">6</span> parcial &middot; <span class="cel c-nao">0</span> sem lançamento &middot;
  <span class="cel c-neu">6</span> 2026 em andamento (esperado).</p>
  <div class="rolagem" id="t-completude"></div>
  <h3 style="margin-top:30px">Lacunas e ressalvas do acervo</h3>
  <div id="t-lacunas"></div>
</div></section>

<section id="acoes"><div class="wrap">
  <span class="tag">Bloco 8 &middot; Encaminhamento</span>
  <h2>Plano de ação</h2>
  <ol class="passos" id="passos"></ol>
</div></section>

<section id="metodo"><div class="wrap">
  <span class="tag">Bloco 9</span>
  <h2>Metodologia, fontes e limites</h2>
  <div id="metodo-txt"></div>
</div></section>

<footer><div class="wrap">
  <p style="margin-bottom:6px"><b>Relatório Consolidado de Concessionárias — BA Ap Log Ex e OMDS</b></p>
  <div class="creditos">
    <img src="assets/e10-mini.jpg" width="198" height="220"
         alt="Emblema do E10 — Gerenciamento de Dados">
    <span>Painel desenvolvido pelo <b>E10 do Comando da BA Ap Log Ex</b><br>
    TC Saldanha &middot; Sgt Rosendo &middot; Sgt De Campos</span>
  </div>
  <p class="sub" style="color:#c8d2c6" id="rodape"></p>
</div></footer>

<script id="dados" type="application/json">__DADOS__</script>
<script>
const D = JSON.parse(document.getElementById('dados').textContent);
const MES = ["jan","fev","mar","abr","mai","jun","jul","ago","set","out","nov","dez"];
const $ = s => document.querySelector(s);

const n = (v, c = 0) => v === null || v === undefined ? '—'
  : v.toLocaleString('pt-BR', {minimumFractionDigits: c, maximumFractionDigits: c});
const rs = (v, c = 2) => 'R$ ' + n(v, c);
const pct = (v, c = 1) => v === null || v === undefined ? '—'
  : (v >= 0 ? '+' : '') + n(v, c) + '%';
const cor = v => v === null ? '' : (v > 0 ? 'pos' : 'neg');
const logo = u => `assets/logos/${u}`;
const NOME = Object.fromEntries(D.ugs.map(u => [u.ug, u.nome]));

/* Célula de identificação da UG, reusada em todas as tabelas. */
const celUG = (ug, sigla, logoArq, extra = '') =>
  `<td title="UG ${ug} — ${NOME[ug]}"><span class="uni">
     <img src="${logo(logoArq)}" alt=""><span><b>${sigla}</b>${extra}<br>
     <span class="sub">UG ${ug}</span></span></span></td>`;

/* ---------------------------------------------------------------- gráficos */
const cssVar = k => getComputedStyle(document.documentElement).getPropertyValue(k).trim();
let PAL, SOBE, DESCE;
function lerTema() {
  PAL = ['--c1','--c2','--c3','--c4','--c5','--c6'].map(cssVar);
  SOBE = cssVar('--sobe'); DESCE = cssVar('--desce');
}
lerTema();

function barras(alvo, itens, {altura = 250, formato = v => n(v), rotulo = ''} = {}) {
  const L = 58, R = 14, T = 16, B = 46, W = 760, H = altura;
  const max = Math.max(...itens.map(i => i.valor), 0) * 1.12 || 1;
  const iw = (W - L - R) / itens.length;
  let g = `<svg viewBox="0 0 ${W} ${H}" style="width:100%;height:auto;display:block" role="img" aria-label="${rotulo}">`;
  for (let k = 0; k <= 4; k++) {
    const y = T + (H - T - B) * k / 4, v = max * (1 - k / 4);
    g += `<line class="grade-h" x1="${L}" y1="${y}" x2="${W - R}" y2="${y}"/>`;
    g += `<text class="rot" x="${L - 8}" y="${y + 4}" text-anchor="end">${formato(v)}</text>`;
  }
  itens.forEach((it, i) => {
    const h = (H - T - B) * (it.valor / max);
    const x = L + i * iw + iw * .18, w = iw * .64, y = H - B - h;
    g += `<rect x="${x}" y="${y}" width="${w}" height="${Math.max(h, 1)}" rx="4" fill="${it.cor || PAL[i % PAL.length]}"/>`;
    g += `<text class="rotb" x="${x + w / 2}" y="${y - 6}" text-anchor="middle">${it.topo || formato(it.valor)}</text>`;
    g += `<text class="rot" x="${x + w / 2}" y="${H - B + 17}" text-anchor="middle">${it.rotulo}</text>`;
    if (it.rotulo2) g += `<text class="rot" x="${x + w / 2}" y="${H - B + 31}" text-anchor="middle">${it.rotulo2}</text>`;
  });
  g += `<line class="eixo" x1="${L}" y1="${H - B}" x2="${W - R}" y2="${H - B}"/></svg>`;
  alvo.innerHTML = g;
}

function barrasDuplas(alvo, itens, serieA, serieB, {altura = 280, formato = v => n(v), rotulo = ''} = {}) {
  const L = 62, R = 122, T = 16, B = 48, W = 760, H = altura;
  const max = Math.max(...itens.flatMap(i => [i.a, i.b])) * 1.14 || 1;
  const iw = (W - L - R) / itens.length;
  let g = `<svg viewBox="0 0 ${W} ${H}" style="width:100%;height:auto;display:block"
    role="img" aria-label="${rotulo || serieA + ' contra ' + serieB}">`;
  for (let k = 0; k <= 4; k++) {
    const y = T + (H - T - B) * k / 4, v = max * (1 - k / 4);
    g += `<line class="grade-h" x1="${L}" y1="${y}" x2="${W - R}" y2="${y}"/>`;
    g += `<text class="rot" x="${L - 8}" y="${y + 4}" text-anchor="end">${formato(v)}</text>`;
  }
  itens.forEach((it, i) => {
    const bw = iw * .30, x0 = L + i * iw + iw * .10;
    [[it.a, PAL[0], 0], [it.b, PAL[3], bw + iw * .06]].forEach(([v, c, dx]) => {
      const h = (H - T - B) * (v / max);
      g += `<rect x="${x0 + dx}" y="${H - B - h}" width="${bw}" height="${Math.max(h, 1)}" rx="3" fill="${c}"/>`;
    });
    g += `<text class="rot" x="${x0 + iw * .33}" y="${H - B + 17}" text-anchor="middle">${it.rotulo}</text>`;
    if (it.nota) g += `<text class="rotb" x="${x0 + iw * .33}" y="${H - B + 32}" text-anchor="middle" fill="${it.notaCor || 'var(--suave)'}">${it.nota}</text>`;
  });
  g += `<rect x="${W - R + 6}" y="${T + 6}" width="11" height="11" rx="2" fill="${PAL[0]}"/>`;
  g += `<text class="rot" x="${W - R + 23}" y="${T + 16}">${serieA}</text>`;
  g += `<rect x="${W - R + 6}" y="${T + 26}" width="11" height="11" rx="2" fill="${PAL[3]}"/>`;
  g += `<text class="rot" x="${W - R + 23}" y="${T + 36}">${serieB}</text>`;
  g += `<line class="eixo" x1="${L}" y1="${H - B}" x2="${W - R}" y2="${H - B}"/></svg>`;
  alvo.innerHTML = g;
}

function cascata(alvo, passos, {altura = 290, rotulo = ''} = {}) {
  const L = 78, R = 14, T = 22, B = 62, W = 760, H = altura;
  let acc = 0; const pts = [];
  passos.forEach(p => {
    if (p.tipo === 'total') { pts.push({...p, ini: 0, fim: p.valor}); acc = p.valor; }
    else { pts.push({...p, ini: acc, fim: acc + p.valor}); acc += p.valor; }
  });
  const vs = pts.flatMap(p => [p.ini, p.fim]);
  const min = Math.min(...vs, 0), max = Math.max(...vs);
  const esc = v => H - B - (H - T - B) * ((v - min) / (max - min || 1));
  const iw = (W - L - R) / pts.length;
  let g = `<svg viewBox="0 0 ${W} ${H}" style="width:100%;height:auto;display:block"
    role="img" aria-label="${rotulo}">`;
  for (let k = 0; k <= 4; k++) {
    const v = min + (max - min) * k / 4, y = esc(v);
    g += `<line class="grade-h" x1="${L}" y1="${y}" x2="${W - R}" y2="${y}"/>`;
    g += `<text class="rot" x="${L - 8}" y="${y + 4}" text-anchor="end">${n(v / 1000)} mil</text>`;
  }
  pts.forEach((p, i) => {
    const x = L + i * iw + iw * .2, w = iw * .6;
    const y1 = esc(p.ini), y2 = esc(p.fim), y = Math.min(y1, y2), h = Math.abs(y2 - y1);
    const c = p.tipo === 'total' ? PAL[0] : (p.valor >= 0 ? SOBE : DESCE);
    g += `<rect x="${x}" y="${y}" width="${w}" height="${Math.max(h, 2)}" rx="3" fill="${c}"/>`;
    g += `<text class="rotb" x="${x + w / 2}" y="${y - 7}" text-anchor="middle">${p.tipo === 'total' ? n(p.valor / 1000) + ' mil' : (p.valor >= 0 ? '+' : '') + n(p.valor / 1000, 1) + ' mil'}</text>`;
    p.rotulo.split('|').forEach((linha, j) =>
      g += `<text class="rot" x="${x + w / 2}" y="${H - B + 17 + j * 13}" text-anchor="middle">${linha}</text>`);
  });
  g += `<line class="eixo" x1="${L}" y1="${esc(0)}" x2="${W - R}" y2="${esc(0)}"/></svg>`;
  alvo.innerHTML = g;
}

/* ------------------------------------------------------------------ conteúdo */
$('#marcas').innerHTML = D.ugs.map(u =>
  `<img src="${logo(u.logo)}" alt="${u.sigla}" title="UG ${u.ug} — ${u.nome} (${u.sigla})" loading="lazy">`).join('');

const T = D.agua.decomposicao.total;
const kwh = D.energia.consumo_total;

$('#resumo-txt').innerHTML = `
<p class="lead">No 1º semestre de 2026 as seis UG apoiadas consumiram
<b>${n(Math.abs(T.var_m), 1)}% menos água</b> que no mesmo período de 2025 — e ainda assim
a despesa com água caiu apenas <b>${n(Math.abs(T.var_r), 1)}%</b>. A diferença tem nome:
a tarifa média efetiva do conjunto passou de <b>${rs(T.p25)}/m³</b> para <b>${rs(T.p26)}/m³</b>
(<b>${pct(T.var_p)}</b>). O aumento de preço custou <b>${rs(T.efeito_preco)}</b> no semestre e
anulou ${n(Math.abs(T.efeito_preco / T.efeito_volume) * 100, 0)}% de toda a economia conquistada
com a redução de consumo.</p>
<p>Na energia elétrica o quadro se inverte: o <b>consumo físico cresceu em ${D.energia.ugs_em_alta} das 6 UG</b>,
mas o valor lançado no SAG <b>não pode ser somado</b> — duas UG lançaram tarifa no lugar do
valor da fatura e uma teve as colunas de ponta e fora ponta invertidas. Este relatório separa
o que os dados provam do que eles apenas sugerem.</p>` + (D.orcamento ? `
<p>O histórico orçamentário fecha a outra ponta e traz <b>a explicação que faltava</b>. O
desembolso com concessionárias subiu
<b>${pct((D.orcamento.caixa_total[D.orcamento.anos[D.orcamento.anos.length - 1]] /
  D.orcamento.caixa_total[D.orcamento.anos[0]] - 1) * 100)}</b> entre
${D.orcamento.anos[0]} e ${D.orcamento.anos[D.orcamento.anos.length - 1]}, e já consome
<b>${n(D.orcamento.peso_total[D.orcamento.anos[D.orcamento.anos.length - 1]].pct, 0)}%</b>
da despesa liquidada das seis UG. Mais decisivo: 2025 liquidou apenas
<b>${n(D.orcamento.rap[D.orcamento.rap.length - 1].liq_sobre_emp, 0)}%</b> do que empenhou e
<b>empurrou ${rs(D.orcamento.rap[D.orcamento.rap.length - 1].levado, 0)} para 2026</b> —
${n(D.orcamento.rap[D.orcamento.rap.length - 1].levado /
    D.orcamento.rap[D.orcamento.rap.length - 2].levado, 1)}× o que o ano anterior havia deixado.</p>`
+ (D.orcamento.corrente ? `
<p>A execução de ${D.orcamento.corrente.rotulo} <b>confirma o diagnóstico</b>:
<b>${n(D.orcamento.corrente.pct_herdado, 1)}%</b> de tudo que as seis UG desembolsaram com
concessionárias este ano foi para quitar contas de exercícios anteriores, e a água está sendo
custeada <b>inteiramente</b> pelo resto a pagar de 2025 — zero empenhado, zero liquidado em 2026.
Parte do aperto que se atribui a 2026 é conta de 2025 chegando atrasada; e o custo próprio do
exercício <b>ainda não apareceu</b>.</p>` : '') : '');

$('#kpis').innerHTML = [
  {r: 'Tarifa média da água', v: rs(T.p26) + '/m³', p: `${pct(T.var_p)} sobre 2025`, c: cor(T.var_p)},
  {r: 'Efeito preço (água)', v: rs(T.efeito_preco, 0), p: 'no 1º semestre de 2026', c: 'pos'},
  {r: 'Consumo de água', v: n(T.m26) + ' m³', p: `${pct(T.var_m)} sobre 2025`, c: cor(T.var_m)},
  {r: 'Consumo de energia', v: n(kwh.k26 / 1000) + ' MWh', p: `${pct(kwh.variacao)} sobre 2025`, c: cor(kwh.variacao)},
  {r: 'UG acima da meta de água', v: `${D.metas.fora_agua} de 6`, p: 'acumulado × meta 2026', c: D.metas.fora_agua > 3 ? 'pos' : ''},
  {r: 'Achados críticos', v: String(D.achados.filter(a => a.grau === 'crítico').length),
   p: 'exigem correção antes do uso', c: 'pos'},
].map(k => `<div class="kpi"><div class="rot">${k.r}</div>
  <div class="num ${k.c}">${k.v}</div><div class="pe">${k.p}</div></div>`).join('');

$('#ressalva').innerHTML = `Dos dois serviços deste relatório, apenas o de <b>água</b> tem valor
financeiro íntegro. Na energia, ${D.energia.tarifarias_sigla} lançaram <b>tarifa unitária</b> no
campo de valor, e a Ba Ap Log Ex está com <b>ponta e fora ponta invertidas</b> em todo o exercício
de 2026. Qualquer número de gasto de energia apresentado adiante é <b>reconstruído</b> e está
marcado como tal — não substitui a conferência contra as faturas originais da concessionária.`;

/* ---- água */
$('#agua-lead').innerHTML = `Entre 2023 e 2026 a tarifa média paga pelo conjunto subiu de
<b>${rs(D.agua.serie[0].tarifa)}/m³</b> para <b>${rs(D.agua.serie[3].tarifa)}/m³</b>
— alta nominal de <b>${n((D.agua.serie[3].tarifa / D.agua.serie[0].tarifa - 1) * 100, 1)}%</b> em
três anos, enquanto o volume consumido no semestre ficou próximo do de 2023
(${n(D.agua.serie[0].m3)} m³ em 2023 contra ${n(D.agua.serie[3].m3)} m³ em 2026).`;

$('#t-agua').innerHTML = `<table><thead><tr>
  <th>Unidade Gestora</th><th class="num">m³ 2025</th><th class="num">m³ 2026</th>
  <th class="num">&Delta; volume</th><th class="num">R$ 2025</th><th class="num">R$ 2026</th>
  <th class="num">&Delta; despesa</th><th class="num">R$/m³ 2026</th><th class="num">&Delta; tarifa</th>
  <th class="num">Efeito preço</th></tr></thead><tbody>` +
  D.agua.decomposicao.itens.map(i => `<tr>
    ${celUG(i.ug, i.sigla, i.logo)}
    <td class="num">${n(i.m25)}</td><td class="num">${n(i.m26)}</td>
    <td class="num ${cor(i.var_m)}">${pct(i.var_m)}</td>
    <td class="num">${n(i.r25, 2)}</td><td class="num">${n(i.r26, 2)}</td>
    <td class="num ${cor(i.var_r)}">${pct(i.var_r)}</td>
    <td class="num">${n(i.p26, 2)}</td>
    <td class="num ${cor(i.var_p)}"><b>${pct(i.var_p)}</b></td>
    <td class="num ${cor(i.efeito_preco)}">${n(i.efeito_preco, 0)}</td></tr>`).join('') +
  `</tbody><tfoot><tr><td>CONJUNTO (6 UG)</td>
    <td class="num">${n(T.m25)}</td><td class="num">${n(T.m26)}</td>
    <td class="num ${cor(T.var_m)}">${pct(T.var_m)}</td>
    <td class="num">${n(T.r25, 2)}</td><td class="num">${n(T.r26, 2)}</td>
    <td class="num ${cor(T.var_r)}">${pct(T.var_r)}</td>
    <td class="num">${n(T.p26, 2)}</td><td class="num ${cor(T.var_p)}">${pct(T.var_p)}</td>
    <td class="num ${cor(T.efeito_preco)}">${n(T.efeito_preco, 0)}</td></tr></tfoot></table>`;

const it = D.agua.decomposicao.itens;
const maisRsMenosM3 = it.filter(i => i.var_r > 0 && i.var_m < 0);
const maisRsMaisM3 = it.filter(i => i.var_r > 0 && i.var_m >= 0);
const tarifaCaiu = it.filter(i => i.var_p < 0);
$('#agua-nota').innerHTML = `<h3 style="margin-top:0">Leitura</h3>
<p>${maisRsMenosM3.length === 1 ? 'Uma UG gastou' : `${maisRsMenosM3.length} UG gastaram`}
<b>mais reais em 2026 consumindo menos água</b>:
${maisRsMenosM3.map(i => `<b>${i.sigla}</b> (${pct(i.var_r)} de despesa com ${pct(i.var_m)} de volume)`).join(' e ')}.
${maisRsMaisM3.length ? `Já ${maisRsMaisM3.map(i => `<b>${i.sigla}</b>`).join(' e ')}
gastou mais com volume praticamente estável (${maisRsMaisM3.map(i => pct(i.var_m)).join(', ')}),
o que também é efeito de tarifa.` : ''}
Nesses casos não há o que corrigir no consumo — a alavanca é contratual e cadastral:
enquadramento tarifário, categoria de uso, percentual de esgoto cobrado e vazamentos não aparentes.</p>
<p style="margin-bottom:0">Na direção oposta,
<b>${tarifaCaiu.map(i => i.sigla).join(', ') || 'nenhuma UG'}</b>
teve queda de tarifa média (${tarifaCaiu.map(i => pct(i.var_p)).join(', ')}), o que merece
verificação: pode indicar correção de cobrança indevida — e, se for o caso, o mesmo erro pode
existir nas demais.</p>`;

/* ---- energia */
$('#energia-lead').innerHTML = `O consumo total das seis UG no 1º semestre foi de
<b>${n(kwh.k26 / 1000)} MWh</b> contra <b>${n(kwh.k25 / 1000)} MWh</b> em 2025
(${pct(kwh.variacao)}), mas o total esconde movimentos opostos:
${D.energia.itens.filter(i => i.variacao > 0).length} UG aumentaram o consumo e
${D.energia.itens.filter(i => i.variacao <= 0).length} reduziram.`;

$('#t-kwh').innerHTML = `<table><thead><tr><th>Unidade Gestora</th><th class="num">Base</th>
  <th class="num">kWh 2023</th><th class="num">kWh 2024</th><th class="num">kWh 2025</th>
  <th class="num">kWh 2026</th><th class="num">&Delta; 25&rarr;26</th></tr></thead><tbody>` +
  D.energia.itens.map(i => `<tr>
    ${celUG(i.ug, i.sigla, i.logo)}
    <td class="num sub">${i.meses} meses</td>
    <td class="num">${n(i.k23)}</td><td class="num">${n(i.k24)}</td>
    <td class="num">${n(i.k25)}</td><td class="num">${n(i.k26)}</td>
    <td class="num ${cor(i.variacao)}"><b>${pct(i.variacao)}</b></td></tr>`).join('') +
  `</tbody><tfoot><tr><td>CONJUNTO (6 UG)</td><td class="num sub">bases mistas</td>
    <td class="num">${n(kwh.k23)}</td><td class="num">${n(kwh.k24)}</td>
    <td class="num">${n(kwh.k25)}</td><td class="num">${n(kwh.k26)}</td>
    <td class="num ${cor(kwh.variacao)}">${pct(kwh.variacao)}</td></tr></tfoot></table>`;

$('#t-rs').innerHTML = `<table><thead><tr><th>Unidade Gestora</th>
  <th class="num">R$ 2023</th><th class="num">R$ 2024</th><th class="num">R$ 2025</th>
  <th class="num">R$ 2026</th><th class="num">R$/kWh 2026</th><th>Situação do dado</th>
  </tr></thead><tbody>` +
  D.energia.itens.map(i => `<tr>
    ${celUG(i.ug, i.sigla, i.logo,
      i.reconstruido ? ' <span class="pilula p-media">R</span>' : '')}
    <td class="num">${n(i.v23, 0)}</td><td class="num">${n(i.v24, 0)}</td>
    <td class="num">${n(i.v25, 0)}</td><td class="num">${n(i.v26, 0)}</td>
    <td class="num">${n(i.v26 / i.k26, 4)}</td>
    <td><span class="pilula ${i.selo_classe}">${i.selo}</span></td></tr>`).join('') +
  `</tbody></table>`;

/* ---- orçamento */
const O = D.orcamento;
if (O) {
  const A = O.anos, u0 = A[0], uN = A[A.length - 1];
  const caixa0 = O.caixa_total[u0], caixaN = O.caixa_total[uN];
  const varCaixa = (caixaN / caixa0 - 1) * 100;
  const rapN = O.rap[O.rap.length - 1], rapAnt = O.rap[O.rap.length - 2];
  const pesoN = O.peso_total[uN], peso0 = O.peso_total[u0];
  const cruzE = O.cruzamento.find(c => c.chave === 'energia');
  const cruzA = O.cruzamento.find(c => c.chave === 'agua');

  $('#orc-lead').innerHTML = `Os crosstabs do Tesouro Gerencial fecham a outra ponta da conta:
  quanto saiu do caixa. Entre ${u0} e ${uN} o desembolso das seis UG com concessionárias passou de
  <b>${rs(caixa0, 0)}</b> para <b>${rs(caixaN, 0)}</b> (<b>${pct(varCaixa)}</b> em
  ${A.length - 1} anos), e o peso desse gasto no custeio subiu de
  <b>${n(peso0.pct, 0)}%</b> para <b>${n(pesoN.pct, 0)}%</b> da despesa liquidada.
  ${O.corrente ? `Os exercícios de ${u0} a ${uN} estão encerrados; o de ${O.corrente.ano} aparece
  em bloco próprio mais abaixo, com a posição acumulada até
  ${O.corrente.mes.toLowerCase()}.` : `Estes arquivos vão até dezembro de ${uN} — <b>não há
  execução do exercício corrente</b>, então este bloco descreve a trajetória que levou até ele.`}`;

  $('#orc-kpis').innerHTML = [
    {r: `Pago em ${uN}`, v: rs(caixaN, 0), p: `${pct(varCaixa)} sobre ${u0}`, c: 'pos'},
    {r: 'Peso no custeio', v: n(pesoN.pct, 0) + '%', p: `da despesa liquidada em ${uN}`,
     c: pesoN.pct > peso0.pct ? 'pos' : 'neg'},
    {r: `Levado de ${uN} para ${+uN + 1}`, v: rs(rapN.levado, 0),
     p: `${n(rapN.levado / rapAnt.levado, 1)}× o do ano anterior`, c: 'pos'},
    {r: 'Liquidado sobre empenhado', v: n(rapN.liq_sobre_emp, 0) + '%', p: `em ${uN}`, c: 'pos'},
    {r: 'Energia — caixa ÷ SAG', v: n(cruzE.razao_media, 2),
     p: `o sistema registra ${n(100 / cruzE.razao_media, 0)}% da fatura`, c: 'pos'},
    {r: 'Água — caixa ÷ SAG', v: n(cruzA.razao_media, 2),
     p: 'perto de 1,00: o dado confere', c: 'neg'},
  ].map(k => `<div class="kpi"><div class="rot">${k.r}</div>
    <div class="num ${k.c}">${k.v}</div><div class="pe">${k.p}</div></div>`).join('');

  const ativos = O.servicos.filter(s => s.ativo);
  $('#t-orc').innerHTML = `<table><thead><tr><th>Plano interno</th>` +
    A.map(a => `<th class="num" colspan="3">${a}</th>`).join('') +
    `</tr><tr><th></th>` + A.map(() =>
      `<th class="num sub">empenhado</th><th class="num sub">liquidado</th><th class="num sub">pago</th>`
    ).join('') + `</tr></thead><tbody>` +
    ativos.map(s => `<tr><td><b>${s.rotulo}</b><br><span class="sub">${s.pi}</span></td>` +
      A.map(a => `<td class="num">${n(s.anos[a].emp, 0)}</td>
                  <td class="num">${n(s.anos[a].liq, 0)}</td>
                  <td class="num"><b>${n(s.anos[a].pago, 0)}</b></td>`).join('') + `</tr>`).join('') +
    `</tbody><tfoot><tr><td>TOTAL</td>` +
    A.map(a => `<td class="num">${n(O.totais[a].emp, 0)}</td>
                <td class="num">${n(O.totais[a].liq, 0)}</td>
                <td class="num">${n(O.caixa_total[a], 0)}</td>`).join('') +
    `</tr></tfoot></table>`;

  $('#t-cruza').innerHTML = `<table><thead><tr><th>Serviço</th>` +
    O.anos_cruzamento.map(a => `<th class="num" colspan="3">${a}</th>`).join('') +
    `</tr><tr><th></th>` + O.anos_cruzamento.map(() =>
      `<th class="num sub">SAG</th><th class="num sub">pago</th><th class="num sub">razão</th>`
    ).join('') + `</tr></thead><tbody>` +
    O.cruzamento.map(c => `<tr><td><b>${c.rotulo}</b></td>` +
      O.anos_cruzamento.map(a => `<td class="num">${n(c.anos[a].sis, 0)}</td>
        <td class="num">${n(c.anos[a].pago, 0)}</td>
        <td class="num ${c.anos[a].razao > 1.05 ? 'pos' : ''}"><b>${n(c.anos[a].razao, 3)}</b></td>`
      ).join('') + `</tr>`).join('') + `</tbody></table>`;

  const disp = [...cruzE.por_ug].sort((a, b) => (b.razao || 0) - (a.razao || 0));
  $('#cruza-nota').innerHTML = `<h3 style="margin-top:0">Leitura</h3>
  <p><b>A água confere.</b> A razão fica em ${O.anos_cruzamento.map(a =>
    n(cruzA.anos[a].razao, 3)).join(', ')} nos três exercícios — o que o SAG registra é o que
  sai do caixa. Isso valida, por fora, toda a análise de tarifa do Bloco 1.</p>
  <p style="margin-bottom:0"><b>A energia não.</b> A razão fica em ${O.anos_cruzamento.map(a =>
    n(cruzE.anos[a].razao, 3)).join(', ')} — estável demais para ser erro pontual. O SAG mede
  consumo faturado em kWh; a fatura cobra também demanda contratada, energia reativa, bandeiras e
  contribuição de iluminação pública. Em ${O.anos_cruzamento[O.anos_cruzamento.length - 1]} a
  dispersão vai de <b>${n(disp[disp.length - 1].razao, 2)}</b>
  (${disp[disp.length - 1].sigla}) a <b>${n(disp[0].razao, 2)}</b> (${disp[0].sigla}),
  o que aponta para diferenças de modalidade tarifária entre as unidades. Ver achado A7.</p>`;

  $('#rap-nota').innerHTML = `<h3 style="margin-top:0">A conta que ${+uN + 1} herdou</h3>
  <p>Em ${uN} foram empenhados <b>${rs(O.totais[uN].emp, 0)}</b> e liquidados
  <b>${rs(O.totais[uN].liq, 0)}</b> — <b>${n(rapN.liq_sobre_emp, 1)}%</b>. A diferença,
  <b>${rs(rapN.levado, 0)}</b>, virou resto a pagar e passou a pressionar o financeiro de
  ${+uN + 1}: é <b>${n(rapN.levado / rapAnt.levado, 1)}×</b> o que ${rapAnt.ano} havia levado
  para ${+rapAnt.ano + 1}.</p>
  <p style="margin-bottom:0" class="sub">A relação foi conferida em todos os exercícios: o
  empenhado menos o liquidado de cada ano reaparece, ao centavo, como resto a pagar não processado
  inscrito no ano seguinte.</p>`;

  /* ---- exercício em curso */
  const C = O.corrente;
  if (C) {
    const herdado = C.pago_total * C.pct_herdado / 100;
    $('#bloco-corrente').innerHTML = `
    <h3 style="margin-top:34px">O exercício em curso — ${C.rotulo}</h3>
    <p class="sub">Posição acumulada do ano até ${C.mes.toLowerCase()}/${C.ano}. Não é comparável
    linha a linha com os exercícios fechados acima: são ${C.n_meses} meses contra doze, e a
    extração usa outro atributo de data.</p>
    <div class="kpis">
      ${[{r: 'Desembolsado no ano', v: rs(C.pago_total, 0), p: `até ${C.mes.toLowerCase()}/${C.ano}`, c: ''},
         {r: 'Disso, conta de anos anteriores', v: n(C.pct_herdado, 1) + '%', p: rs(herdado, 0), c: 'pos'},
         {r: 'Liquidado do próprio exercício', v: rs(C.liquidado, 0), p: `de ${rs(C.empenhado, 0)} empenhados`, c: 'neg'},
         {r: 'Resto a pagar já quitado', v: n(C.rap_pct_quitado, 0) + '%', p: `restam ${rs(C.rap_saldo, 0)}`, c: ''},
        ].map(k => `<div class="kpi"><div class="rot">${k.r}</div>
          <div class="num ${k.c}">${k.v}</div><div class="pe">${k.p}</div></div>`).join('')}
    </div>
    <div class="rolagem"><table><thead><tr><th>Plano interno</th>
      <th class="num">Empenhado ${C.ano}</th><th class="num">Liquidado ${C.ano}</th>
      <th class="num">RAP herdado</th><th class="num">RAP quitado</th>
      <th class="num">Desembolso total</th></tr></thead><tbody>` +
      C.servicos.map(s => `<tr><td><b>${s.rotulo}</b><br><span class="sub">${s.pi}</span></td>
        <td class="num">${n(s.emp, 2)}</td><td class="num">${n(s.liq, 2)}</td>
        <td class="num">${n(s.rap_herdado, 2)}</td><td class="num">${n(s.rap_pago, 2)}</td>
        <td class="num"><b>${n(s.pago, 2)}</b></td></tr>`).join('') +
      `</tbody><tfoot><tr><td>TOTAL</td>
        <td class="num">${n(C.empenhado, 2)}</td><td class="num">${n(C.liquidado, 2)}</td>
        <td class="num">${n(C.rap_herdado, 2)}</td><td class="num">${n(C.rap_quitado, 2)}</td>
        <td class="num">${n(C.pago_total, 2)}</td></tr></tfoot></table></div>
    <div class="cartao perigo"><h3 style="margin-top:0">O que ${C.ano} está pagando de verdade</h3>
      <p><b>${n(C.pct_herdado, 1)}% do desembolso de ${C.ano}</b> (${rs(herdado, 0)} de
      ${rs(C.pago_total, 0)}) foi para quitar restos a pagar de exercícios anteriores. A despesa
      do próprio exercício soma apenas ${rs(C.composicao[0].v, 0)}
      (${n(C.composicao[0].pct, 1)}%).</p>
      ${(() => { const ag = C.servicos.find(s => s.chave === 'agua');
        return ag && !ag.emp && !ag.liq ? `<p><b>A água é o caso extremo:</b> zero empenhado e
        zero liquidado em ${C.ano}, com ${rs(ag.rap_pago, 0)} pagos contra o resto a pagar
        herdado. As seis UG estão sendo abastecidas com crédito de
        ${+C.ano - 1}.</p>` : ''; })()}
      <p style="margin-bottom:0">O estoque herdado é de ${rs(C.rap_herdado, 2)} —
      ${rs(C.rap_inscritos, 2)} inscritos ao fim de ${+C.ano - 1} mais
      ${rs(C.rap_reinscritos, 2)} reinscritos de exercícios mais antigos. Desse total,
      <b>${n(C.rap_pct_quitado, 0)}%</b> já foi quitado e restam ${rs(C.rap_saldo, 0)}.
      Quando o crédito herdado acabar, a despesa corrente passa a aparecer inteira no exercício —
      é aí que o aperto de ${C.ano} fica visível no orçamento.</p></div>`;
  }

  const tel = O.servicos.filter(s => ['telm', 'telf', 'corr'].includes(s.chave));
  $('#t-tel').innerHTML = `<table><thead><tr><th>Plano interno</th>` +
    A.map(a => `<th class="num">${a}</th>`).join('') +
    `<th class="num">${u0}&rarr;${uN}</th></tr></thead><tbody>` +
    tel.map(s => `<tr><td><b>${s.rotulo}</b><br><span class="sub">${s.pi}</span></td>` +
      A.map(a => `<td class="num">${n(s.anos[a].pago, 2)}</td>`).join('') +
      `<td class="num ${cor(s.var_pago)}">${pct(s.var_pago, 0)}</td></tr>`).join('') +
    `</tbody></table>`;

  const telm = O.servicos.find(s => s.chave === 'telm');
  const telf = O.servicos.find(s => s.chave === 'telf');
  const corr = O.servicos.find(s => s.chave === 'corr');
  const fatiaTel = (telm.anos[uN].pago + telf.anos[uN].pago + corr.anos[uN].pago) / caixaN * 100;
  $('#tel-nota').innerHTML = `<h3 style="margin-top:0">Escala</h3>
  <p>Telefonia e correios somados são <b>${n(fatiaTel, 2)}%</b> do desembolso com concessionárias
  em ${uN}. A <b>telefonia móvel (PTELM)</b> pagou ${rs(telm.anos[uN].pago, 2)} no ano nas seis UG
  — ordem de ${rs(telm.anos[uN].pago / 6 / 12, 0)} por unidade por mês. É pouco dinheiro, e por
  isso mesmo não é onde está o problema de 2026.</p>
  <p style="margin-bottom:0">A <b>telefonia fixa</b> saiu de cena: o último pagamento relevante
  foi em ${A.filter(a => telf.anos[a].pago > 1000).pop() || u0}
  (${rs(telf.anos[A.filter(a => telf.anos[a].pago > 1000).pop() || u0].pago, 2)}), e o plano está
  zerado desde ${A.find(a => telf.anos[a].pago === 0) || uN}. Se ainda existe linha fixa em uso,
  ela está sendo paga por outro plano interno — vale confirmar onde.</p>`;

  $('#t-peso').innerHTML = `<table><thead><tr><th>Unidade Gestora</th>` +
    A.map(a => `<th class="num">${a}</th>`).join('') + `</tr></thead><tbody>` +
    O.peso.map(p => `<tr>${celUG(p.ug, p.sigla, p.logo)}` +
      A.map(a => `<td class="num">${p.anos[a].pct === null ? '—' : n(p.anos[a].pct, 0) + '%'}
        <br><span class="sub">${n(p.anos[a].conc / 1000, 0)} mil</span></td>`).join('') +
      `</tr>`).join('') +
    `</tbody><tfoot><tr><td>CONJUNTO</td>` +
    A.map(a => `<td class="num">${n(O.peso_total[a].pct, 0)}%<br>
      <span class="sub">${n(O.peso_total[a].conc / 1000, 0)} mil</span></td>`).join('') +
    `</tr></tfoot></table>`;

  const anomalo = O.peso.flatMap(p => A.filter(a =>
    p.anos[a].pct !== null && p.anos[a].pct < 5).map(a => ({sigla: p.sigla, ano: a, ...p.anos[a]})));
  $('#peso-nota').innerHTML = `O percentual é a fatia das concessionárias na despesa liquidada
  total da UG, e por isso <b>cai quando a unidade executa uma despesa grande e atípica</b> no ano.
  ${anomalo.length ? anomalo.map(x => `É o caso de <b>${x.sigla} em ${x.ano}</b>, cujo total
  liquidado foi de ${rs(x.total / 1000000, 1)} milhões — o gasto com concessionárias
  (${rs(x.conc / 1000, 0)} mil) não caiu, o denominador é que inchou.`).join(' ') : ''}
  A tendência de fundo é clara: no conjunto, a fatia subiu de ${n(O.peso_total[u0].pct, 0)}%
  para ${n(O.peso_total[uN].pct, 0)}% entre ${u0} e ${uN}.`;
} else {
  document.getElementById('orcamento').style.display = 'none';
  document.querySelector('nav.indice a[href="#orcamento"]').closest('li').style.display = 'none';
}

/* ---- achados */
const GRAU ={'crítico': ['p-alta', 'Crítico'], 'alto': ['p-media', 'Alto'],
  'médio': ['p-media', 'Médio'], 'informativo': ['p-info', 'Informativo']};
$('#lista-achados').innerHTML = D.achados.map(a => {
  const [cls, rot] = GRAU[a.grau];
  return `<details class="achado" ${a.grau === 'crítico' ? 'open' : ''}>
    <summary><span class="pilula ${cls}">${rot}</span>
      <span><b>${a.id} &middot; ${a.titulo}</b><br>
      <span class="sub">${a.servico} &middot; ${a.sigla}</span></span></summary>
    <div class="corpo">
      <p>${a.resumo}</p>
      <span class="rot-mini">Evidência</span><ul>${a.evidencias.map(e => `<li>${e}</li>`).join('')}</ul>
      <span class="rot-mini">Impacto</span><p>${a.impacto}</p>
      <span class="rot-mini">Encaminhamento</span><p style="margin-bottom:0">${a.acao}</p>
    </div></details>`;
}).join('');

/* ---- cards por UG */
$('#cards-ug').innerHTML = D.ugs.map(u => {
  const ag = D.agua.decomposicao.itens.find(i => i.ug === u.ug);
  const en = D.energia.itens.find(i => i.ug === u.ug);
  const ma = D.metas.agua.find(i => i.ug === u.ug);
  return `<div class="card-ug">
    <header><img src="${logo(u.logo)}" alt="">
      <div><h3>${u.sigla}</h3>
        <div class="nomeug">${u.nome}</div>
        <div class="ugid">UG ${u.ug} &middot; 1ª RM &middot; RJ</div></div></header>
    <div class="mini">
      <div><div class="r">Água — tarifa</div><div class="v">${rs(ag.p26)}</div>
        <div class="sub ${cor(ag.var_p)}">${pct(ag.var_p)} vs 2025</div></div>
      <div><div class="r">Água — despesa</div><div class="v">R$ ${n(ag.r26 / 1000)} mil</div>
        <div class="sub ${cor(ag.var_r)}">${pct(ag.var_r)} vs 2025</div></div>
      <div><div class="r">Energia — consumo</div><div class="v">${n(en.k26 / 1000)} MWh</div>
        <div class="sub ${cor(en.variacao)}">${pct(en.variacao)} vs 2025</div></div>
      <div><div class="r">Água × meta</div><div class="v ${ma.desvio > 0 ? 'pos' : 'neg'}">${pct(ma.desvio)}</div>
        <div class="sub">acum. até ${MES[ma.mes - 1]}/2026</div></div>
    </div>
    ${u.alertas.length ? `<div style="margin-top:12px">${u.alertas.map(a =>
      `<div class="sub" style="margin-bottom:4px"><span class="pilula p-alta">${a.id}</span> ${a.txt}</div>`).join('')}</div>` : ''}
  </div>`;
}).join('');

/* ---- metas */
$('#metas-lead').innerHTML = `A meta lançada no SAG é uma fração fixa da média histórica que o
próprio sistema calcula, mês a mês — apurado em todos os registros dos arquivos:
<b>85% para água</b>, <b>85% para energia fora ponta</b> e <b>90% para energia ponta</b>.
Barras acima de 100% indicam UG acima da meta, isto é, consumindo mais do que o alvo.
<b>Atenção:</b> essa média ignora os anos sem lançamento, então nem sempre é trienal — no fora
ponta do ECT ela se apoia em <b>um único ano</b> (2025), pois 2023 e 2024 estão zerados.`;

const em = D.metas.energia;
$('#t-meta').innerHTML = `<table><thead><tr><th>Unidade Gestora</th><th class="num">Base</th>
  <th class="num">Ponta acum.</th><th class="num">Meta ponta</th><th class="num">&Delta;</th>
  <th class="num">Fora ponta acum.</th><th class="num">Meta fora ponta</th><th class="num">&Delta;</th>
  <th class="num">Meta do ano (kWh)</th></tr></thead><tbody>` +
  em.map(m => `<tr>
    ${celUG(m.ug, m.sigla, m.logo)}
    <td class="num sub">até ${MES[m.mes - 1]}</td>
    <td class="num">${n(m.ponta_ac)}</td><td class="num">${n(m.ponta_meta)}</td>
    <td class="num ${cor(m.ponta_desvio)}">${pct(m.ponta_desvio)}</td>
    <td class="num">${n(m.fora_ac)}</td><td class="num">${n(m.fora_meta)}</td>
    <td class="num ${cor(m.fora_desvio)}">${pct(m.fora_desvio)}</td>
    <td class="num">${n(m.ponta_meta_ano + m.fora_meta_ano)}</td></tr>`).join('') +
  `</tbody></table>`;

$('#meta-nota').innerHTML = `<h3 style="margin-top:0">Três ressalvas de leitura</h3>
<ul style="margin-bottom:0">
<li><b>Meses de corte diferentes.</b> BMSA e ECT param em mai/2026 na energia, o BCMS já acumula
até jul/2026 nos dois serviços, as demais até jun. As barras <b>não são comparáveis entre si</b>
sem esse ajuste (achado A5).</li>
<li><b>Ba Ap Log Ex, energia.</b> A soma ponta + fora ponta não é afetada pela inversão do
achado A1 — mas a abertura por posto sim: a UG aparece com ${pct(em[0].ponta_desvio, 0)} na ponta
e ${pct(em[0].fora_desvio, 0)} fora dela. Os dois números estão <b>trocados entre si</b>, não são
desempenho real.</li>
<li><b>ECT, energia.</b> Os ${pct(em[4].desvio)} de desvio total são ilusórios: a meta ainda
reserva ${n(em[4].ponta_meta)} kWh para a ponta, onde a UG deixou de lançar desde 2025
(achado A3). Só no fora ponta, o ECT está ${pct(em[4].fora_desvio)} em relação à meta.</li>
</ul>`;

/* ---- acervo */
$('#acervo-lead').innerHTML = `O acervo tem <b>${D.acervo.total} arquivos do SAG</b>
reorganizados em <b>${D.acervo.pastas} pastas</b> por concessionária e unidade, mais
<b>${D.acervo.orcamentarios} crosstabs do Tesouro Gerencial</b> com a execução orçamentária.
As seis UG têm o mesmo conjunto no SAG: <b>6 séries de energia</b> (kWh e R$ em ponta e fora
ponta, mais consumo × meta nos dois postos) e <b>3 séries de água</b> (m³, R$ e consumo × meta),
além dos dois relatórios consolidados de meta. <b>A simetria entre as unidades está completa.</b>`;

$('#t-completude').innerHTML = `<table class="matriz"><thead><tr><th>UG</th><th>Série</th>` +
  ['2023', '2024', '2025', '2026'].map(a => `<th class="num">${a}</th>`).join('') +
  `<th>Ressalvas em 2023–2025</th></tr></thead><tbody>` +
  D.completude.map(l => {
    const anos = ['2023', '2024', '2025'];
    const semSerie = anos.filter(a => l.anos[a].n === 0);
    const parciais = anos.filter(a => l.anos[a].n > 0 && l.anos[a].faltando.length)
      .map(a => `${a}: sem ${l.anos[a].faltando.join(', ')}`);
    const obs = [];
    if (semSerie.length) obs.push(`<b>sem lançamento em ${semSerie.join(' e ')}</b>`);
    if (parciais.length) obs.push(parciais.join(' · '));
    return `<tr><td><b>${l.sigla}</b></td><td>${l.metrica}</td>` +
      ['2023', '2024', '2025', '2026'].map(a => {
        const q = l.anos[a].n;
        const c = a === '2026' ? 'c-neu' : (q === 12 ? 'c-ok' : (q === 0 ? 'c-nao' : 'c-par'));
        return `<td class="num"><span class="cel ${c}">${q}</span></td>`;
      }).join('') +
      `<td class="sub">${obs.join(' · ') || '—'}</td></tr>`;
  }).join('') + `</tbody></table>`;

$('#t-lacunas').innerHTML = `<div class="cartao ok"><h3 style="margin-top:0">${D.lacunas.resolvido.titulo}</h3>
  <p style="margin-bottom:0">${D.lacunas.resolvido.txt}</p></div>
  <div class="cartao aviso"><h3 style="margin-top:0">Pendências</h3>
  <ul style="margin-bottom:0">${D.lacunas.itens.map(i => `<li>${i}</li>`).join('')}</ul></div>`;

/* ---- ações */
$('#passos').innerHTML = D.acoes.map(a =>
  `<li><b>${a.t}</b><br><span class="sub">${a.d}</span></li>`).join('');

/* ---- gráficos (redesenhados quando o tema muda) */
function desenharGraficos() {
  lerTema();

  barras($('#g-tarifa'), D.agua.serie.map((s, i) => ({
    rotulo: s.ano, rotulo2: `${n(s.m3)} m³`, valor: s.tarifa, topo: rs(s.tarifa),
    cor: i === 3 ? SOBE : PAL[0],
  })), {formato: v => n(v, 1), altura: 270,
    rotulo: 'Tarifa média de água paga pelo conjunto, por ano, jan a junho'});

  cascata($('#g-cascata'), [
    {rotulo: 'Despesa|jan–jun 2025', valor: T.r25, tipo: 'total'},
    {rotulo: 'Efeito volume|(' + pct(T.var_m) + ' de m³)', valor: T.efeito_volume},
    {rotulo: 'Efeito preço|(' + pct(T.var_p) + ' na tarifa)', valor: T.efeito_preco},
    {rotulo: 'Efeito cruzado', valor: T.efeito_misto},
    {rotulo: 'Despesa|jan–jun 2026', valor: T.r26, tipo: 'total'},
  ], {rotulo: 'Decomposição da variação da despesa com água entre 2025 e 2026 em efeito volume, efeito preço e efeito cruzado'});

  barrasDuplas($('#g-kwh'),
    D.energia.itens.map(i => ({rotulo: i.sigla, a: i.k25 / 1000, b: i.k26 / 1000,
      nota: pct(i.variacao), notaCor: i.variacao > 0 ? SOBE : DESCE})),
    'jan–jun 2025 (MWh)', 'jan–jun 2026 (MWh)',
    {formato: v => n(v, 0),
     rotulo: 'Consumo de energia por UG, comparando o 1º semestre de 2025 com o de 2026'});

  barras($('#g-meta-agua'), D.metas.agua.map(m => ({
    rotulo: m.sigla, rotulo2: `${n(m.acumulado)} m³ até ${MES[m.mes - 1]}`,
    valor: (m.acumulado / m.meta) * 100, topo: pct(m.desvio),
    cor: m.desvio > 0 ? SOBE : DESCE,
  })), {formato: v => n(v, 0) + '%', altura: 275,
    rotulo: 'Água: acumulado sobre a meta, por UG'});

  barras($('#g-meta-fora'), em.map(m => ({
    rotulo: m.sigla, rotulo2: `${n(m.fora_ac / 1000)} MWh até ${MES[m.mes - 1]}`,
    valor: (m.fora_ac / m.fora_meta) * 100, topo: pct(m.fora_desvio),
    cor: ['160238', '160321'].includes(m.ug) ? PAL[2] : (m.fora_desvio > 0 ? SOBE : DESCE),
  })), {formato: v => n(v, 0) + '%', altura: 275,
    rotulo: 'Energia fora ponta: acumulado sobre a meta, por UG'});

  barras($('#g-meta-ponta'), em.map(m => ({
    rotulo: m.sigla, rotulo2: `${n(m.ponta_ac / 1000)} MWh até ${MES[m.mes - 1]}`,
    valor: Math.min((m.ponta_ac / m.ponta_meta) * 100, 200), topo: pct(m.ponta_desvio, 0),
    cor: ['160238', '160321'].includes(m.ug) ? PAL[2] : (m.ponta_desvio > 0 ? SOBE : DESCE),
  })), {formato: v => n(v, 0) + '%', altura: 275,
    rotulo: 'Energia ponta: acumulado sobre a meta, por UG'});

  $('#ponta-legenda').innerHTML = `A barra da Ba Ap Log Ex está <b>truncada em 200%</b> para não
  achatar as demais — o valor real é ${pct(em[0].ponta_desvio, 0)} e é artefato do achado A1,
  não consumo. As barras em cinza-oliva marcam as duas UG cujo dado de posto tarifário está
  comprometido.`;

  if (O) {
    barras($('#g-caixa'), O.anos.map((a, i) => ({
      rotulo: a, rotulo2: rs(O.caixa_total[a] / 1000, 0) + ' mil',
      valor: O.caixa_total[a], topo: n(O.caixa_total[a] / 1000000, 2) + ' mi',
      cor: i === O.anos.length - 1 ? SOBE : PAL[0],
    })), {formato: v => n(v / 1000000, 1) + ' mi', altura: 260,
      rotulo: 'Pagamentos totais do ano com concessionárias'});

    barras($('#g-rap'), O.rap.map(r => ({
      rotulo: `${r.ano} → ${+r.ano + 1}`,
      rotulo2: `liq. ${n(r.liq_sobre_emp, 0)}% do empenho`,
      valor: r.levado, topo: n(r.levado / 1000000, 2) + ' mi',
      cor: r.ano === O.anos[O.anos.length - 1] ? SOBE : PAL[1],
    })), {formato: v => n(v / 1000000, 1) + ' mi', altura: 260,
      rotulo: 'Restos a pagar levados de um exercício ao seguinte'});
  }
}
desenharGraficos();
matchMedia('(prefers-color-scheme:dark)').addEventListener('change', desenharGraficos);

/* ---- metodologia e rodapé */
$('#metodo-txt').innerHTML = D.metodo;
$('#rodape').innerHTML = D.rodape;

/* ---- navegação ativa */
const links = [...document.querySelectorAll('nav.indice a')];
const obs = new IntersectionObserver(es => es.forEach(e => {
  if (e.isIntersecting) links.forEach(l =>
    l.classList.toggle('on', l.getAttribute('href') === '#' + e.target.id));
}), {rootMargin: '-45% 0px -50% 0px'});
document.querySelectorAll('section[id]').forEach(s => obs.observe(s));
</script>
</body>
</html>
"""


# ------------------------------------------------------------------------ montagem

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--fonte", default=FONTE_PADRAO)
    args = ap.parse_args()

    series, meta_agua, meta_energia, consolidado, inventario, n_pastas = coletar(args.fonte)
    orc_bruto, arq_orc = coletar_orcamento(args.fonte)
    tarifarias = detecta_tarifario(series)
    dec = decomposicao_agua(series)
    ser_agua = serie_tarifa_agua(series)
    bases = bases_energia(series)
    val_energia = energia_valores(series, tarifarias, bases)
    ms = metas(consolidado, meta_energia)
    comp = completude(series)
    orc = orcamento(orc_bruto, series, tarifarias)
    ach = achados(series, tarifarias, dec, ms, ser_agua, orc)

    itens_e, tot = [], {"k23": 0, "k24": 0, "k25": 0, "k26": 0}
    for u in UGS:
        ug, s = u["ug"], series[u["ug"]]
        nm = bases[ug]
        k = {a: soma(s["kwh_ponta"][a], nm) + soma(s["kwh_fora"][a], nm) for a in ANOS}
        v = val_energia[ug]
        if ug == "160238":
            selo, cls = "Colunas invertidas (A1)", "p-alta"
        elif v["reconstruido"]:
            selo, cls = "Reconstruído de tarifa (A2)", "p-media"
        elif ug == "160321":
            selo, cls = "Posto tarifário trocado (A3)", "p-media"
        elif ug == "160304":
            selo, cls = "Atípicos em 2025 (A4)", "p-media"
        else:
            selo, cls = "Consistente", "p-baixa"
        itens_e.append({
            "ug": ug, "sigla": u["sigla"], "logo": u["logo"], "meses": nm,
            "k23": k["2023"], "k24": k["2024"], "k25": k["2025"], "k26": k["2026"],
            "variacao": var(k["2025"], k["2026"]),
            "v23": v["valores"]["2023"], "v24": v["valores"]["2024"],
            "v25": v["valores"]["2025"], "v26": v["valores"]["2026"],
            "reconstruido": v["reconstruido"], "selo": selo, "selo_classe": cls,
        })
        for a, kk in (("k23", "2023"), ("k24", "2024"), ("k25", "2025"), ("k26", "2026")):
            tot[a] += k[kk]
    tot["variacao"] = var(tot["k25"], tot["k26"])

    alertas = {u["ug"]: [] for u in UGS}
    for a in ach:
        for ug in a["ug"].split("/"):
            if ug in alertas and a["grau"] in ("crítico", "alto"):
                alertas[ug].append({"id": a["id"], "txt": a["titulo"]})
    ugs_pub = [dict(u, alertas=alertas[u["ug"]]) for u in UGS]

    dados = {
        "gerado_em": datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
        "ugs": ugs_pub,
        "agua": {"decomposicao": dec, "serie": ser_agua},
        "energia": {
            "itens": itens_e, "consumo_total": tot,
            "ugs_em_alta": sum(1 for i in itens_e if i["variacao"] and i["variacao"] > 0),
            "tarifarias": tarifarias,
            "tarifarias_sigla": " e ".join(UGIDX[t]["sigla"] for t in tarifarias),
        },
        "metas": ms,
        "achados": ach,
        "completude": comp,
        "orcamento": orc,
        "acervo": {"total": len(inventario), "pastas": n_pastas,
                   "orcamentarios": len(arq_orc), "arquivos_orc": arq_orc},
        "lacunas": {
            "resolvido": {
                "titulo": "Lacuna fechada — acompanhamento de meta da energia",
                "txt": "As seis UG passaram a ter o gráfico de <b>consumo × meta de energia</b> "
                       "nos dois postos tarifários (ponta e fora ponta), equivalente ao que a água "
                       "já tinha. São 12 arquivos que <b>completam a simetria do acervo</b> e "
                       "permitiram, nesta versão, separar a meta de ponta da meta de fora ponta — "
                       "o que revelou que as duas seguem percentuais diferentes.",
            },
            "itens": [
                "<b>ECT — energia 2023/2024 sem fora ponta e 2025/2026 sem ponta.</b> Não é "
                "arquivo faltando: é a série que muda de posto tarifário no meio (achado A3).",
                "<b>ECT — nov e dez/2024 sem valor lançado</b> (R$ 0,00 com consumo registrado).",
                "<b>BMSA — nov/2025 de água zerado</b> em consumo e valor: fatura não lançada.",
                "<b>BMSA e ECT — energia sem jun/2026</b>; o BCMS já tem jul/2026. Mês de corte "
                "diferente entre UG (achado A5).",
                "<b>Os arquivos não identificam a concessionária.</b> O SAG exporta UG, "
                "sigla, RM, comando, CGCFEx e UF — nunca o nome da distribuidora, a matrícula "
                "da instalação nem a denominação por extenso da unidade. A organização “por "
                "concessionária” foi feita por <i>serviço</i> (energia elétrica / água e esgoto), "
                "e os nomes das UG neste relatório foram informados pela BA Ap Log Ex.",
                "<b>Não há faturas originais no acervo.</b> Nenhum documento da concessionária "
                "permite confirmar os achados A1 a A4 nem decompor os cerca de 16% da fatura de "
                "energia que o SAG não registra (achado A7).",
                "<b>O exercício de 2026 está em curso.</b> O crosstab traz a posição acumulada "
                "até agosto e usa outro atributo de data (“Mês Lançamento”) — os valores de 2026 "
                "não são comparáveis linha a linha com os exercícios encerrados. O que se pode "
                "afirmar com segurança é a composição do desembolso e o abatimento do resto a "
                "pagar, que fecham entre si.",
            ],
        },
        "acoes": [
            {"t": "Corrigir a inversão ponta/fora ponta da Ba Ap Log Ex em 2026 (A1)",
             "d": "Reprocessar os seis meses no SAG conferindo contra as faturas da "
                  "distribuidora. Até lá, nenhum indicador de energia por posto tarifário da "
                  "UG 160238 deve ser usado para decisão ou prestação de contas."},
            {"t": "Padronizar o campo VALOR do SAG (A2)",
             "d": "Orientar 1º D Sup e BCMS a lançar o valor total da fatura, com tributos e "
                  "bandeira tarifária, e recompor a série 2023–2026 dessas duas UG."},
            {"t": "Reunir as faturas originais do 1º semestre de 2026",
             "d": "Seis meses × seis UG × dois serviços = 72 faturas. Sem elas nenhum achado "
                  "deste relatório pode virar pedido de revisão junto à concessionária."},
            {"t": "Abrir revisão cadastral e tarifária da água nas UG com alta de tarifa",
             "d": "Ba Ap Log Ex, DC Mun, 1º D Sup, ECT e BCMS tiveram alta de R$/m³. Verificar "
                  "categoria de uso, faixas de consumo, percentual de esgoto cobrado e "
                  "hidrômetros com submedição."},
            {"t": "Investigar as quedas atípicas de tarifa antes de comemorar",
             "d": "Onde a tarifa média caiu de forma acentuada, confirmar se houve mudança de "
                  "modalidade, correção de cobrança indevida ou apenas fatura não lançada."},
            {"t": "Fixar mês de corte único para o relatório consolidado (A5)",
             "d": "Consolidar apenas UG fechadas até o mês de referência e sinalizar as demais, "
                  "em vez de somar bases desiguais."},
            {"t": "Acompanhar mensalmente o esgotamento do resto a pagar de 2025 (A8)",
             "d": "Enquanto o saldo herdado cobrir as faturas, a despesa corrente não aparece no "
                  "exercício. Quando acabar, o custo de 2026 surge de uma vez — e é preciso ter "
                  "crédito empenhado antes disso, sobretudo na água, que hoje roda inteiramente "
                  "sobre o saldo de 2025."},
            {"t": "Dimensionar o empenho pela liquidação histórica (A8)",
             "d": "Em 2025 o empenho das concessionárias foi quase o dobro do liquidado, e a "
                  "diferença virou resto a pagar. Empenho estimativo calibrado reduz o estoque "
                  "de RAP sem risco de inadimplência."},
            {"t": "Decompor uma fatura completa de energia por UG (A7)",
             "d": "O SAG capta cerca de 84% do que se paga em energia. Identificar quanto é "
                  "demanda contratada, reativo, bandeira e iluminação pública mostra onde há "
                  "economia sem reduzir consumo — sobretudo nas UG com razão mais alta."},
            {"t": "Usar a meta por posto tarifário, não a meta total",
             "d": "A meta de ponta é 90% da média trienal e a de fora ponta é 85%. Somar as duas "
                  "e comparar com o consumo total mistura dois critérios e esconde desvios, como "
                  "ficou evidente no ECT."},
        ],
        "metodo": """
<h3>Fonte</h3>
<p>Arquivos exportados do <b>SAG — Sistema de Acompanhamento da Gestão</b>, cobrindo jan/2023 a
jul/2026, das seis UG apoiadas
pela BA Ap Log Ex — todas em <b>1ª RM / CML / 1º CGCFEx / RJ</b>, conforme os próprios arquivos
informam. São as séries históricas por UG e serviço, os gráficos de consumo × meta de 2026 e
dois relatórios consolidados de meta. Os arquivos têm extensão <code>.xls</code>, mas o conteúdo
é HTML de tabela do Highcharts; foram lidos diretamente do HTML, sem Excel.</p>
<p>A execução orçamentária vem dos <b>crosstabs do Tesouro Gerencial</b> (“CRÉDITO DISP”),
cobrindo os exercícios de <b>2022 a 2025</b> encerrados e a posição parcial de <b>2026 até
agosto</b>, nos Planos Internos das concessionárias:
<code>I3DACSPAGES</code> (água e esgoto), <code>I3DACSPENEL</code> (energia e iluminação pública),
<code>I3DACSPTELM</code> (telefonia móvel), <code>I3DACSPTELF</code> (telefonia fixa) e
<code>I3DACSPCORR</code> (postais). Os crosstabs trazem linhas de subtotal por Plano Interno;
elas são descartadas na leitura — sem isso o valor dobraria. As colunas de natureza 339000
carregam apenas provisão e crédito disponível, e por isso não entram nos totais de despesa.
O layout do cabeçalho muda entre extrações (às vezes a métrica vem antes do período, às vezes
depois), então as duas linhas são localizadas pelo conteúdo. <b>O arquivo de 2026 usa o atributo
“Mês Lançamento” em vez de “NC — Dia Emissão”</b>, por isso a posição parcial do exercício é
apresentada em bloco separado, e não na mesma série dos anos fechados.</p>

<p>Dois dados <b>não vêm de nenhuma das duas fontes</b> e precisam de registro à parte:</p>
<ul>
<li>A <b>denominação por extenso das unidades</b> foi informada pela BA Ap Log Ex; o sistema
exporta apenas a sigla. Todo o resto da identificação (UG, RM, comando, CGCFEx, UF) vem
dos arquivos.</li>
<li>O <b>nome da concessionária</b> continua sem fonte: não aparece em nenhum arquivo do acervo,
nem há matrícula ou número de instalação. Por isso a organização “por concessionária” foi feita
por <i>serviço</i>.</li>
</ul>

<h3>Recortes</h3>
<ul>
<li><b>Água:</b> jan–jun de cada ano, seis meses fechados nas seis UG.</li>
<li><b>Energia:</b> mesma janela, reduzida a jan–mai nas UG sem jun/2026 lançado
(BMSA e ECT). A coluna “Base” da tabela de energia informa o recorte de cada UG; o total do
conjunto soma bases diferentes e está assinalado como tal.</li>
<li>Jul/2026 do BCMS foi excluído de todas as comparações entre anos.</li>
</ul>

<h3>Fórmulas</h3>
<ul>
<li><b>Tarifa implícita</b> = valor lançado &divide; consumo lançado. Não é a tarifa homologada da
concessionária; é o preço efetivamente pago por unidade, já com tributos, bandeiras e encargos
embutidos, como chega ao orçamento.</li>
<li><b>Decomposição preço/volume:</b> efeito preço = (p₂₆ &minus; p₂₅) &times; q₂₅;
efeito volume = (q₂₆ &minus; q₂₅) &times; p₂₅;
efeito cruzado = (p₂₆ &minus; p₂₅) &times; (q₂₆ &minus; q₂₅). A soma dos três reproduz exatamente
a variação da despesa — o que serve de conferência do cálculo.</li>
<li><b>Gasto de energia reconstruído</b> = &Sigma; (kWh do mês &times; tarifa lançada no mês),
aplicado apenas às UG do achado A2.</li>
<li><b>Meta:</b> percentual fixo sobre a média histórica calculada pelo próprio SAG, mês a mês.
Verificado em todos os registros dos arquivos: 85% na água, 85% na energia fora ponta e 90% na
energia ponta. Essa média <b>desconsidera os anos sem lançamento</b> — não é necessariamente
trienal, e no fora ponta do ECT reduz-se a um único ano.</li>
</ul>

<h3>Limites</h3>
<ul>
<li>Os achados A1 a A4 são inferidos da <b>consistência interna dos próprios dados</b>. São
fortes o bastante para suspender o uso dos números, mas <b>só a fatura original confirma</b> o que
aconteceu.</li>
<li>Onde a tarifa implícita de energia cai de forma atípica em 2026, há mais de uma explicação
possível — entre elas mudança de modalidade tarifária ou faturamento parcial. <b>Este relatório
não afirma qual é o caso</b>: o acervo não permite decidir.</li>
<li>Nenhum valor foi corrigido pela inflação. Todas as variações são <b>nominais</b>.</li>
<li>A projeção anual do efeito preço da água supõe que o 2º semestre repita o comportamento do
1º; é ordem de grandeza, não previsão.</li>
</ul>""",
        "rodape": "Fonte: SAG (Sistema de Acompanhamento da Gestão) — relatórios de consumo e "
                  "meta de água e energia — e Tesouro Gerencial, para a execução orçamentária. UG "
                  "160238, 160246, 160304, 160307, 160321 e 160329 (1ª RM / CML / 1º CGCFEx / RJ). "
                  "Séries de jan/2023 a jul/2026. Página estática, sem coleta de dados do "
                  "visitante. Documento de trabalho para uso interno: os achados A1 a A4 devem ser "
                  "conferidos contra as faturas originais antes de qualquer providência externa.",
    }

    os.makedirs(SAIDA, exist_ok=True)
    with open(os.path.join(SAIDA, "dados.json"), "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=1)
    with open(os.path.join(SAIDA, "index.html"), "w", encoding="utf-8") as f:
        f.write(gerar_html(dados))
    open(os.path.join(SAIDA, ".nojekyll"), "w").close()

    print(f"OK  {len(inventario)} arquivos lidos em {n_pastas} pastas")
    print(f"    valor = tarifa em: {[UGIDX[t]['sigla'] for t in tarifarias]}")
    print(f"    água   tarifa {dec['total']['p25']:.2f} -> {dec['total']['p26']:.2f} "
          f"({dec['total']['var_p']:+.1f}%) | efeito preço R$ {dec['total']['efeito_preco']:+,.2f}")
    print(f"    energia kWh {tot['k25']:,.0f} -> {tot['k26']:,.0f} ({tot['variacao']:+.1f}%)")
    if orc:
        a0, aN = orc["anos"][0], orc["anos"][-1]
        r = orc["rap"][-1]
        print(f"    orçamento {len(arq_orc)} arquivos, {a0}–{aN}: caixa "
              f"R$ {orc['caixa_total'][a0]:,.2f} -> R$ {orc['caixa_total'][aN]:,.2f}")
        print(f"      RAP levado de {aN} para {int(aN)+1}: R$ {r['levado']:,.2f} "
              f"(liquidou {r['liq_sobre_emp']:.1f}% do empenhado)")
        for c in orc["cruzamento"]:
            print(f"      razão caixa/SAG {c['chave']}: "
                  + ", ".join(f"{x}={c['anos'][x]['razao']:.3f}" for x in orc["anos_cruzamento"]))
        cr = orc["corrente"]
        if cr:
            print(f"      {cr['rotulo']}: desembolso R$ {cr['pago_total']:,.2f}, "
                  f"{cr['pct_herdado']:.1f}% quitando exercícios anteriores "
                  f"(composição fecha: {cr['confere']})")
            print(f"        RAP herdado R$ {cr['rap_herdado']:,.2f} -> quitado "
                  f"{cr['rap_pct_quitado']:.1f}%, saldo R$ {cr['rap_saldo']:,.2f}")
    print(f"    site -> {os.path.join(SAIDA, 'index.html')}")


if __name__ == "__main__":
    main()
